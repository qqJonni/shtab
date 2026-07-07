"""
Детерминированный парсер смет для xlsx и csv.
ИИ-фолбэк для PDF и нераспознанных форматов — заглушка (шаг 4).

Алгоритм:
  1. Ищем строку-заголовок в первых 30 строках: ≥2 распознанных столбца включая «name».
  2. Маппим столбцы по синонимам (см. _HEADER_SYNONYMS).
  3. Парсим строки ниже: нормализуем числа (рус. формат) и единицы.
  4. Отфильтровываем мусор: пустые строки, итоги, разделы, заголовки групп.
  5. Считаем confidence по наполненности + арифметической консистентности.

Возвращает list[dict]:
  {name, unit, quantity, unit_price, total, confidence,
   unit_unknown (bool), raw{unit, quantity, unit_price, total}}
"""

import re
import csv
import json
from typing import Optional

from openpyxl import load_workbook

import config


# ── Синонимы заголовков колонок ────────────────────────────────────────────
_HEADER_SYNONYMS: dict[str, list[str]] = {
    'name': [
        'наименование работ и затрат', 'наименование работ', 'наименование (вид работ)',
        'наименование позиции', 'наименование', 'вид работ', 'виды работ',
        'состав работ', 'описание работ', 'описание', 'работы', 'наим.', 'наим', 'name',
    ],
    'unit': [
        'единица измерения', 'ед. измерения', 'ед.изм.', 'ед.изм', 'ед.', 'ед',
        'единица', 'unit',
    ],
    'quantity': [
        'количество (объем)', 'количество/объем', 'количество (объём)',
        'объём работ', 'объем работ', 'количество', 'объём', 'объем',
        'кол-во', 'кол.', 'кол', 'qty',
        # «Всего» / «Общий объём» в сметах — суммарное кол-во, не сумма денег
        'всего', 'общий объем', 'общий объём', 'итого объем', 'итого объём',
        'общее количество', 'объем по всем', 'объём по всем',
    ],
    'unit_price': [
        'стоимость за единицу', 'стоимость за ед.', 'стоимость за ед',
        'цена за единицу', 'цена за ед.', 'цена за ед',
        'стоимость единицы', 'стоимость ед.', 'стоимость ед',
        'ед. цена', 'цена ед.', 'расценка за ед.', 'расценка', 'цена', 'price',
    ],
    'total': [
        'общая стоимость', 'итого стоимость', 'общая сумма', 'стоимость работ',
        'сумма работ', 'итого, руб', 'итого руб', 'итого', 'стоимость',
        'сумма', 'total',
    ],
}

# ── Маппинг единиц → config.UNITS ─────────────────────────────────────────
_UNIT_MAP: dict[str, str] = {
    # м2
    'м2': 'м2', 'м²': 'м2', 'кв.м.': 'м2', 'кв.м': 'м2', 'кв м': 'м2',
    'кв.метр': 'м2', 'кв.метров': 'м2', 'кв метр': 'м2',
    # м.пог.
    'м.пог.': 'м.пог.', 'м.пог': 'м.пог.', 'пог.м.': 'м.пог.', 'пог.м': 'м.пог.',
    'п.м.': 'м.пог.', 'п.м': 'м.пог.', 'пм': 'м.пог.',
    'м.п.': 'м.пог.', 'м.п': 'м.пог.',    # сокращение в ряде смет
    'м. пог.': 'м.пог.', 'м. пог': 'м.пог.',  # с пробелом после «м»
    # шт.
    'шт.': 'шт.', 'шт': 'шт.', 'штук': 'шт.', 'штуки': 'шт.', 'штука': 'шт.',
    # м3
    'м3': 'м3', 'м³': 'м3', 'куб.м.': 'м3', 'куб.м': 'м3', 'куб м': 'м3',
    'м.куб.': 'м3', 'м.куб': 'м3', 'куб.метр': 'м3', 'куб.метров': 'м3',
    # м.
    'м.': 'м.', 'м': 'м.',
    # ч.
    'ч.': 'ч.', 'ч': 'ч.', 'час': 'ч.', 'часов': 'ч.', 'чел.ч': 'ч.', 'чел.ч.': 'ч.',
    # т.
    'т.': 'т.', 'т': 'т.', 'тн': 'т.', 'тонн': 'т.', 'тонна': 'т.', 'тонн.': 'т.',
    # кг.
    'кг.': 'кг.', 'кг': 'кг.',
}

# Паттерны строк-разделов и итогов
_JUNK_RE = re.compile(
    r'^\s*(итого|всего|в\s*том\s*числе|в\s*т\.?ч\.?|раздел|глава|часть|'
    r'примечание|note|прим\.?|№\s*п/?п|n\s*п/?п|наименование|наим\.)\s*[:\.]?\s*$',
    re.IGNORECASE,
)
_ROMAN_RE = re.compile(r'^\s*[IVXLCDM]+[\.\s\)]+', re.IGNORECASE)


# ── Утилиты ────────────────────────────────────────────────────────────────

def _normalize_num(val) -> Optional[float]:
    """Конвертирует «1 234,56» / «1234.56» / int/float → float. None если не число."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s in ('-', '—', 'x', 'х', '*'):
        return None
    s = re.sub(r'\s', '', s)          # убрать пробелы-разделители тысяч
    s = s.replace(',', '.')           # русская десятичная запятая → точка
    # «1.234.567» — несколько точек: слить всё кроме последней
    parts = s.split('.')
    if len(parts) > 2:
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        v = float(s)
        return v if v >= 0 else None  # отрицательные суммы не берём
    except ValueError:
        return None


def _normalize_unit(raw) -> tuple[str, bool]:
    """Возвращает (нормализованная единица, known). known=False → нужен ручной выбор."""
    if not raw:
        return '', False
    key = str(raw).strip().lower()
    if key in _UNIT_MAP:
        return _UNIT_MAP[key], True
    # Уже правильное значение из config.UNITS
    stripped = str(raw).strip()
    if stripped in config.UNITS:
        return stripped, True
    return stripped, False


def _match_header(cell) -> Optional[tuple[str, int]]:
    """
    Возвращает (field, score) или None.
    score = длина совпавшего синонима — длиннее = точнее.
    """
    if cell is None:
        return None
    t = str(cell).strip().lower()
    t = re.sub(r'\s*\n\s*', ' ', t)          # многострочные заголовки → одна строка
    t = re.sub(r'[,\s]+руб[\.ьёя]?.*$', '', t).strip()
    t = re.sub(r'\s+', ' ', t)
    best: Optional[tuple[str, int]] = None
    for field, synonyms in _HEADER_SYNONYMS.items():
        for syn in synonyms:
            if t == syn or t.startswith(syn + ' ') or t.startswith(syn + ','):
                if best is None or len(syn) > best[1]:
                    best = (field, len(syn))
    return best


def _is_junk(name: str, nums: list) -> bool:
    """True → строку пропускаем."""
    if not name or not name.strip():
        return True
    n = name.strip()
    if len(n) < 3:
        return True
    if _JUNK_RE.match(n):
        return True
    if _ROMAN_RE.match(n) and all(v is None for v in nums):
        return True  # «I. Земляные работы» без цифр — раздел
    if all(v is None for v in nums):
        return True  # строка только с текстом — скорее всего заголовок группы
    return False


def _confidence(row: dict) -> float:
    score = 0.0
    if row.get('name'):
        score += 0.30
    if row.get('unit'):
        score += 0.10
    if not row.get('unit_unknown'):
        score += 0.10  # единица распознана
    if row.get('quantity') is not None:
        score += 0.20
    if row.get('unit_price') is not None:
        score += 0.15
    if row.get('total') is not None:
        score += 0.15
    # Арифметическая консистентность: total ≈ quantity × unit_price
    q = row.get('quantity')
    p = row.get('unit_price')
    t = row.get('total')
    if q and p and t and t > 0:
        if abs(q * p - t) / t < 0.02:
            score = min(score + 0.05, 1.0)
    return round(min(score, 1.0), 2)


# ── Общая логика поиска заголовка и парсинга строк ────────────────────────

def _parse_rows(raw_rows: list[list]) -> list[dict]:
    """
    Принимает список строк (каждая — список значений ячеек).
    Возвращает список распознанных позиций.
    """
    if not raw_rows:
        return []

    col_map: dict[str, int] = {}
    header_idx = -1

    # Ищем строку-заголовок.
    # Среди нескольких совпадений для одного поля побеждает наиболее длинный синоним
    # (например «общий объем» бьёт «объем» для колонки 'quantity').
    for i, row in enumerate(raw_rows[:30]):
        scores: dict[str, tuple[int, int]] = {}  # field -> (score, col_idx)
        for j, cell in enumerate(row):
            result_h = _match_header(cell)
            if result_h:
                field, score = result_h
                if field not in scores or score > scores[field][0]:
                    scores[field] = (score, j)
        mapping = {f: idx for f, (_, idx) in scores.items()}
        if 'name' in mapping and len(mapping) >= 2:
            col_map = mapping
            header_idx = i
            break

    if not col_map or 'name' not in col_map:
        return []  # таблицу не нашли

    result = []
    for raw_row in raw_rows[header_idx + 1:]:

        def get(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(raw_row):
                return None
            v = raw_row[idx]
            return v if v != '' else None

        name_cell = get('name')
        name = str(name_cell).strip() if name_cell is not None else ''

        quantity  = _normalize_num(get('quantity'))
        unit_price = _normalize_num(get('unit_price'))
        total     = _normalize_num(get('total'))

        if _is_junk(name, [quantity, unit_price, total]):
            continue

        raw_unit = str(get('unit') or '').strip()
        unit, unit_known = _normalize_unit(raw_unit)

        row_dict = {
            'name':         name,
            'unit':         unit,
            'unit_unknown': not unit_known,
            'quantity':     quantity,
            'unit_price':   unit_price,
            'total':        total,
            'raw': {
                'unit':       raw_unit,
                'quantity':   str(get('quantity') or ''),
                'unit_price': str(get('unit_price') or ''),
                'total':      str(get('total') or ''),
            },
        }
        row_dict['confidence'] = _confidence(row_dict)
        result.append(row_dict)

    return result


# ── Форматно-специфичные парсеры ───────────────────────────────────────────

def parse_xlsx(filepath: str) -> list[dict]:
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = [
        list(row)
        for row in ws.iter_rows(values_only=True)
        if any(c is not None for c in row)
    ]
    return _parse_rows(rows)


def parse_csv(filepath: str) -> list[dict]:
    for enc in ('utf-8-sig', 'cp1251', 'utf-8'):
        try:
            with open(filepath, newline='', encoding=enc) as f:
                sample = f.read(8192)
            with open(filepath, newline='', encoding=enc) as f:
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
                except csv.Error:
                    dialect = csv.excel  # fallback: запятая
                reader = csv.reader(f, dialect)
                rows = [row for row in reader if any(c.strip() for c in row)]
            return _parse_rows(rows)
        except (UnicodeDecodeError, csv.Error):
            continue
    return []


# ── Публичная точка входа ─────────────────────────────────────────────────

def parse_file(filepath: str, source_type: str) -> list[dict]:
    """
    Возвращает список позиций (может быть пустым — тогда статус 'failed').
    source_type: 'xlsx' | 'csv' | 'pdf'
    """
    try:
        if source_type == 'xlsx':
            return parse_xlsx(filepath)
        elif source_type == 'csv':
            return parse_csv(filepath)
        else:
            return []  # PDF → ИИ-фолбэк будет добавлен на шаге 4
    except Exception:
        return []
