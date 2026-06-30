import io
from flask import render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

from db import query_db
from helpers import role_required


def register(app):

    @app.route('/reports')
    @login_required
    @role_required('manager', 'admin', 'pto')
    def reports_page():
        objects = query_db("SELECT id, name FROM objects WHERE status='active' ORDER BY name")
        contractors = query_db("SELECT id, name FROM organizations WHERE type='contractor' ORDER BY name")
        return render_template('reports/index.html', objects=objects, contractors=contractors)

    @app.route('/api/report-data')
    @login_required
    @role_required('manager', 'admin', 'pto')
    def api_report_data():
        report_type = request.args.get('report', 'substages_status')
        obj_id = request.args.get('object_id', '')
        stage_id = request.args.get('stage_id', '')
        contractor_id = request.args.get('contractor_id', '')

        where_obj = f'AND cs.object_id = {int(obj_id)}' if obj_id else ''
        where_stage = f'AND cs.id = {int(stage_id)}' if stage_id else ''
        where_con = f'AND cs.contractor_id = {int(contractor_id)}' if contractor_id else ''
        filters = f'{where_obj} {where_stage} {where_con}'

        if report_type == 'substages_status':
            rows = query_db(
                f"SELECT ss.status, COUNT(*) as cnt FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"WHERE 1=1 {filters} GROUP BY ss.status")
            labels_map = {'not_started': 'Не начат', 'in_progress': 'В работе', 'done': 'Выполнен',
                          'closed': 'Закрыт', 'approved': 'Согласован'}
            colors_map = {'not_started': '#94A3B8', 'in_progress': '#3B82F6', 'done': '#F59E0B',
                          'closed': '#8B5CF6', 'approved': '#10B981'}
            return jsonify({
                'title': 'Подэтапы по статусам',
                'labels': [labels_map.get(r['status'], r['status']) for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [colors_map.get(r['status'], '#94A3B8') for r in rows],
            })

        elif report_type == 'stages_status':
            rows = query_db(
                f"SELECT cs.status, COUNT(*) as cnt FROM construction_stages cs "
                f"WHERE 1=1 {where_obj} {where_con} GROUP BY cs.status")
            labels_map = {'planned': 'Планируется', 'in_progress': 'В работе', 'done': 'Завершён', 'suspended': 'Приостановлен'}
            colors_map = {'planned': '#94A3B8', 'in_progress': '#3B82F6', 'done': '#10B981', 'suspended': '#F59E0B'}
            return jsonify({
                'title': 'Этапы по статусам',
                'labels': [labels_map.get(r['status'], r['status']) for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [colors_map.get(r['status'], '#94A3B8') for r in rows],
            })

        elif report_type == 'defects_priority':
            rows = query_db(
                f"SELECT d.priority, COUNT(*) as cnt FROM defects d "
                f"JOIN construction_stages cs ON d.stage_id = cs.id "
                f"WHERE d.status NOT IN ('closed','verified') {filters} GROUP BY d.priority")
            labels_map = {'low': 'Низкий', 'normal': 'Обычный', 'high': 'Высокий', 'critical': 'Критический'}
            colors_map = {'low': '#94A3B8', 'normal': '#3B82F6', 'high': '#F59E0B', 'critical': '#EF4444'}
            return jsonify({
                'title': 'Замечания по приоритетам (открытые)',
                'labels': [labels_map.get(r['priority'], r['priority']) for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [colors_map.get(r['priority'], '#94A3B8') for r in rows],
            })

        elif report_type == 'defects_status':
            rows = query_db(
                f"SELECT d.status, COUNT(*) as cnt FROM defects d "
                f"JOIN construction_stages cs ON d.stage_id = cs.id "
                f"WHERE 1=1 {filters} GROUP BY d.status")
            labels_map = {'open': 'Открыто', 'in_progress': 'В работе', 'resolved': 'Устранено',
                          'verified': 'Проверено', 'rejected': 'Отклонено', 'closed': 'Закрыто'}
            colors_map = {'open': '#EF4444', 'in_progress': '#3B82F6', 'resolved': '#F59E0B',
                          'verified': '#10B981', 'rejected': '#94A3B8', 'closed': '#10B981'}
            return jsonify({
                'title': 'Замечания по статусам',
                'labels': [labels_map.get(r['status'], r['status']) for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [colors_map.get(r['status'], '#94A3B8') for r in rows],
            })

        elif report_type == 'substages_by_object':
            rows = query_db(
                f"SELECT o.name, COUNT(ss.id) as cnt FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"JOIN objects o ON cs.object_id = o.id "
                f"WHERE 1=1 {where_con} GROUP BY o.id ORDER BY o.name")
            return jsonify({
                'title': 'Подэтапы по объектам',
                'labels': [r['name'] for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': ['#3B82F6'] * len(rows),
            })

        elif report_type == 'substages_by_contractor':
            rows = query_db(
                f"SELECT COALESCE(org.name, 'Не назначен') as name, COUNT(ss.id) as cnt "
                f"FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"LEFT JOIN organizations org ON cs.contractor_id = org.id "
                f"WHERE 1=1 {where_obj} GROUP BY cs.contractor_id ORDER BY name")
            palette = ['#3B82F6', '#10B981', '#F59E0B', '#EF4444', '#8B5CF6', '#EC4899', '#94A3B8']
            return jsonify({
                'title': 'Подэтапы по подрядчикам',
                'labels': [r['name'] for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [palette[i % len(palette)] for i in range(len(rows))],
            })

        elif report_type == 'defects_by_contractor':
            rows = query_db(
                f"SELECT COALESCE(org.name, 'Не назначен') as name, COUNT(d.id) as cnt "
                f"FROM defects d "
                f"LEFT JOIN organizations org ON d.contractor_id = org.id "
                f"JOIN construction_stages cs ON d.stage_id = cs.id "
                f"WHERE 1=1 {where_obj} GROUP BY d.contractor_id ORDER BY cnt DESC")
            palette = ['#EF4444', '#F59E0B', '#3B82F6', '#10B981', '#8B5CF6', '#94A3B8']
            return jsonify({
                'title': 'Замечания по подрядчикам',
                'labels': [r['name'] for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [palette[i % len(palette)] for i in range(len(rows))],
            })

        elif report_type == 'substages_overdue_by_object':
            rows = query_db(
                f"SELECT o.name, COUNT(ss.id) as cnt FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"JOIN objects o ON cs.object_id = o.id "
                f"WHERE ss.plan_end_date < to_char(now(),'YYYY-MM-DD') AND ss.status NOT IN ('done','closed','approved') "
                f"{where_con} GROUP BY o.id ORDER BY cnt DESC")
            palette = ['#EF4444', '#F59E0B', '#3B82F6', '#8B5CF6', '#94A3B8']
            return jsonify({
                'title': 'Просроченные подэтапы по объектам',
                'labels': [r['name'] for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [palette[i % len(palette)] for i in range(len(rows))],
            })

        elif report_type == 'substages_overdue_by_contractor':
            rows = query_db(
                f"SELECT COALESCE(org.name, 'Не назначен') as name, COUNT(ss.id) as cnt "
                f"FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"LEFT JOIN organizations org ON cs.contractor_id = org.id "
                f"WHERE ss.plan_end_date < to_char(now(),'YYYY-MM-DD') AND ss.status NOT IN ('done','closed','approved') "
                f"{where_obj} GROUP BY cs.contractor_id ORDER BY cnt DESC")
            palette = ['#EF4444', '#F59E0B', '#3B82F6', '#8B5CF6', '#94A3B8']
            return jsonify({
                'title': 'Просроченные подэтапы по подрядчикам',
                'labels': [r['name'] for r in rows],
                'data': [r['cnt'] for r in rows],
                'colors': [palette[i % len(palette)] for i in range(len(rows))],
            })

        elif report_type == 'table_substages_overdue':
            rows = query_db(
                f"SELECT o.name as obj_name, cs.name as stage_name, ss.name, ss.volume, ss.unit, "
                f"ss.status, ss.plan_end_date, COALESCE(org.name, '') as contractor_name "
                f"FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"JOIN objects o ON cs.object_id = o.id "
                f"LEFT JOIN organizations org ON cs.contractor_id = org.id "
                f"WHERE ss.plan_end_date < to_char(now(),'YYYY-MM-DD') AND ss.status NOT IN ('done','closed','approved') "
                f"{filters} ORDER BY ss.plan_end_date")
            status_map = {'not_started': 'Не начат', 'in_progress': 'В работе'}
            return jsonify({
                'title': 'Просроченные подэтапы (таблица)',
                'headers': ['Объект', 'Этап', 'Подэтап', 'Объём', 'Ед.', 'Статус', 'Срок', 'Подрядчик'],
                'rows': [[r['obj_name'], r['stage_name'], r['name'],
                          r['volume'] or '', r['unit'] or '',
                          status_map.get(r['status'], r['status']),
                          r['plan_end_date'] or '', r['contractor_name']] for r in rows],
            })

        elif report_type == 'table_substages':
            rows = query_db(
                f"SELECT o.name as obj_name, cs.name as stage_name, ss.name, ss.volume, ss.unit, "
                f"ss.status, ss.plan_end_date, COALESCE(org.name, '') as contractor_name "
                f"FROM substages ss "
                f"JOIN construction_stages cs ON ss.stage_id = cs.id "
                f"JOIN objects o ON cs.object_id = o.id "
                f"LEFT JOIN organizations org ON cs.contractor_id = org.id "
                f"WHERE 1=1 {filters} ORDER BY o.name, cs.order_num, ss.id")
            status_map = {'not_started': 'Не начат', 'in_progress': 'В работе', 'done': 'Выполнен',
                          'closed': 'Закрыт', 'approved': 'Согласован'}
            return jsonify({
                'title': 'Подэтапы (таблица)',
                'headers': ['Объект', 'Этап', 'Подэтап', 'Объём', 'Ед.', 'Статус', 'Срок', 'Подрядчик'],
                'rows': [[r['obj_name'], r['stage_name'], r['name'],
                          r['volume'] or '', r['unit'] or '',
                          status_map.get(r['status'], r['status']),
                          r['plan_end_date'] or '', r['contractor_name']] for r in rows],
            })

        elif report_type == 'table_defects':
            rows = query_db(
                f"SELECT d.id, o.name as obj_name, cs.name as stage_name, d.title, d.priority, "
                f"d.status, d.due_date, d.reopen_count, COALESCE(org.name, '') as contractor_name, "
                f"u.full_name as reporter "
                f"FROM defects d "
                f"JOIN objects o ON d.object_id = o.id "
                f"JOIN construction_stages cs ON d.stage_id = cs.id "
                f"LEFT JOIN organizations org ON d.contractor_id = org.id "
                f"LEFT JOIN users u ON d.reporter_id = u.id "
                f"WHERE 1=1 {filters} ORDER BY d.created_at DESC")
            pri_map = {'low': 'Низкий', 'normal': 'Обычный', 'high': 'Высокий', 'critical': 'Критический'}
            st_map = {'open': 'Открыто', 'in_progress': 'В работе', 'resolved': 'Устранено',
                      'verified': 'Проверено', 'rejected': 'Отклонено', 'closed': 'Закрыто'}
            return jsonify({
                'title': 'Замечания (таблица)',
                'headers': ['#', 'Объект', 'Этап', 'Замечание', 'Приоритет', 'Статус', 'Срок', 'Возвраты', 'Подрядчик'],
                'rows': [[r['id'], r['obj_name'], r['stage_name'], r['title'],
                          pri_map.get(r['priority'], ''), st_map.get(r['status'], ''),
                          r['due_date'] or '', r['reopen_count'], r['contractor_name']] for r in rows],
            })

        return jsonify({'title': 'Нет данных', 'labels': [], 'data': [], 'colors': []})

    @app.route('/api/report-export')
    @login_required
    @role_required('manager', 'admin', 'pto')
    def api_report_export():
        # Get data by calling the view function directly
        resp = api_report_data()
        data = resp.get_json()

        wb = Workbook()
        ws = wb.active
        ws.title = (data.get('title', 'Отчёт'))[:31]

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hf = Font(bold=True, size=10)
        hfill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws.cell(1, 1, data.get('title', 'Отчёт')).font = Font(bold=True, size=13)

        if data.get('headers'):
            # Table export
            headers = data['headers']
            for i, h in enumerate(headers, 1):
                c = ws.cell(3, i, h)
                c.font = hf; c.fill = hfill; c.border = border
                c.alignment = Alignment(horizontal='center', wrap_text=True)
            for ri, row in enumerate(data.get('rows', []), 4):
                for ci, val in enumerate(row, 1):
                    c = ws.cell(ri, ci, val)
                    c.border = border
        elif data.get('labels'):
            # Chart data as table
            ws.cell(3, 1, 'Показатель').font = hf
            ws.cell(3, 1).fill = hfill; ws.cell(3, 1).border = border
            ws.cell(3, 2, 'Значение').font = hf
            ws.cell(3, 2).fill = hfill; ws.cell(3, 2).border = border
            for i, (label, val) in enumerate(zip(data['labels'], data['data']), 4):
                ws.cell(i, 1, label).border = border
                ws.cell(i, 2, val).border = border

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        for col in 'CDEFGHI':
            ws.column_dimensions[col].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_title = data.get('title', 'Отчёт').replace('/', '-')[:50]
        return send_file(buf, as_attachment=True,
                         download_name=f'{safe_title}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
