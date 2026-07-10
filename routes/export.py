import io
from flask import send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from db import query_db
from flask_login import current_user
from helpers import role_required
from reports import objects_summary


def _style_sheet(ws, title, headers):
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = Font(bold=True, size=10)
    hfill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(1, 1, title).font = Font(bold=True, size=13)

    for i, h in enumerate(headers, 1):
        c = ws.cell(3, i, h)
        c.font = hf
        c.fill = hfill
        c.border = border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    return border


def register(app):

    @app.route('/export/summary')
    @login_required
    @role_required('manager', 'admin', 'pto')
    def export_summary():
        wb = Workbook()

        # ═══ Лист 1: Объекты ═══
        ws = wb.active
        ws.title = 'Объекты'
        headers = ['Объект', 'Адрес', '% готовности', 'Этапов', 'Подэтапов', 'Выполнено',
                   'Замечания (откр.)', 'Пакеты (акт.)', 'Сумма завершённых (₽)']
        border = _style_sheet(ws, 'Сводка по объектам', headers)

        for cw, w in [(1,25),(2,25),(3,12),(4,8),(5,10),(6,10),(7,14),(8,12),(9,20)]:
            ws.column_dimensions[chr(64+cw)].width = w

        dev_id = current_user.organization_id if current_user.role != 'admin' else None
        objs = objects_summary(dev_id)
        for row_i, o in enumerate(objs, 4):
            vals = [o['name'], o.get('address',''), o['progress'],
                    o['stages_count'], o['substages_total'], o['substages_done'],
                    o['defects_open'], o['packages_active'], o['completed_sum']]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row_i, col, val)
                c.border = border
                if isinstance(val, float):
                    c.number_format = '#,##0.00'

        # ═══ Лист 2: Этапы/Подэтапы ═══
        ws2 = wb.create_sheet('Этапы и подэтапы')
        headers2 = ['Объект', 'Этап', 'Подрядчик', 'Статус этапа', 'Подэтап',
                     'Объём', 'Ед.', 'Расценка (₽)', 'Сумма (₽)', 'Статус', 'Срок']
        border2 = _style_sheet(ws2, 'Этапы и подэтапы', headers2)
        for cw, w in [(1,20),(2,20),(3,18),(4,14),(5,25),(6,10),(7,8),(8,12),(9,14),(10,12),(11,12)]:
            ws2.column_dimensions[chr(64+cw)].width = w

        tenant_where = f' AND o.developer_id = {current_user.organization_id}' if (current_user.role != 'admin' and current_user.organization_id) else ''
        stages = query_db(
            f'SELECT cs.*, o.name as obj_name, org.name as contractor_name '
            f'FROM construction_stages cs '
            f'JOIN objects o ON cs.object_id=o.id '
            f'LEFT JOIN organizations org ON cs.contractor_id=org.id '
            f'WHERE 1=1{tenant_where} ORDER BY o.name, cs.order_num')
        status_map = {'planned':'Планируется','in_progress':'В работе','done':'Завершён','suspended':'Приостановлен',
                      'not_started':'Не начат','closed':'Закрыт','approved':'Согласован'}
        r = 4
        for s in stages:
            subs = query_db('SELECT * FROM substages WHERE stage_id=? ORDER BY id', (s['id'],))
            if not subs:
                vals = [s['obj_name'], s['name'], s['contractor_name'] or '', status_map.get(s['status'],''),
                        '', '', '', '', '', '', '']
                for col, val in enumerate(vals, 1):
                    ws2.cell(r, col, val).border = border2
                r += 1
            else:
                for i, sub in enumerate(subs):
                    vals = [s['obj_name'] if i==0 else '', s['name'] if i==0 else '',
                            (s['contractor_name'] or '') if i==0 else '',
                            status_map.get(s['status'],'') if i==0 else '',
                            sub['name'], sub['volume'], sub['unit'] or '',
                            sub['unit_price'], sub['total_price'],
                            status_map.get(sub['status'],''), sub['plan_end_date'] or '']
                    for col, val in enumerate(vals, 1):
                        c = ws2.cell(r, col, val)
                        c.border = border2
                        if isinstance(val, float) and col in (8,9):
                            c.number_format = '#,##0.00'
                    r += 1

        # ═══ Лист 3: Замечания ═══
        ws3 = wb.create_sheet('Замечания')
        headers3 = ['#', 'Объект', 'Этап', 'Тип', 'Приоритет', 'Статус',
                     'Автор', 'Подрядчик', 'Срок', 'Возвраты', 'Создано']
        border3 = _style_sheet(ws3, 'Замечания', headers3)
        for cw, w in [(1,5),(2,20),(3,18),(4,12),(5,12),(6,12),(7,18),(8,18),(9,12),(10,8),(11,12)]:
            ws3.column_dimensions[chr(64+cw)].width = w

        pri_map = {'low':'Низкий','normal':'Обычный','high':'Высокий','critical':'Критический'}
        st_map = {'open':'Открыто','in_progress':'В работе','resolved':'Устранено',
                  'verified':'Проверено','rejected':'Отклонено','closed':'Закрыто'}
        defects = query_db(
            'SELECT d.*, o.name as obj_name, cs.name as stage_name, dt.name as type_name, '
            'u.full_name as reporter_name, org.name as contractor_name '
            'FROM defects d JOIN objects o ON d.object_id=o.id '
            'JOIN construction_stages cs ON d.stage_id=cs.id '
            'LEFT JOIN defect_types dt ON d.type_id=dt.id '
            'LEFT JOIN users u ON d.reporter_id=u.id '
            'LEFT JOIN organizations org ON d.contractor_id=org.id '
            'ORDER BY d.created_at DESC')
        for row_i, d in enumerate(defects, 4):
            vals = [d['id'], d['obj_name'], d['stage_name'], d['type_name'] or '',
                    pri_map.get(d['priority'],''), st_map.get(d['status'],''),
                    d['reporter_name'] or '', d['contractor_name'] or '',
                    d['due_date'] or '', d['reopen_count'],
                    (d['created_at'] or '')[:10]]
            for col, val in enumerate(vals, 1):
                ws3.cell(row_i, col, val).border = border3

        # ═══ Лист 4: Согласования ═══
        ws4 = wb.create_sheet('Согласования')
        headers4 = ['Пакет #', 'Подэтап', 'Объект', 'Этап', 'Подрядчик', 'Статус',
                     'Технадзор', 'Прораб', 'ПТО', 'Руководитель', 'Бухгалтерия', 'Отправлен']
        border4 = _style_sheet(ws4, 'Согласования пакетов', headers4)
        for cw, w in [(1,8),(2,20),(3,18),(4,18),(5,18),(6,14),(7,12),(8,12),(9,12),(10,12),(11,12),(12,12)]:
            ws4.column_dimensions[chr(64+cw)].width = w

        pkg_st = {'draft':'Черновик','in_review':'На согласовании','returned':'Возвращён',
                  'approved':'Согласован','completed':'Завершён'}
        step_st = {'waiting':'⏳','pending':'🔵','approved':'✅','returned':'❌'}
        packages = query_db(
            'SELECT dp.*, '
            "(SELECT string_agg(ss.name, '; ') FROM package_items pi "
            ' JOIN substages ss ON pi.substage_id = ss.id WHERE pi.package_id = dp.id) as sub_name, '
            'cs.name as stage_name, '
            'o.name as obj_name, org.name as con_name '
            'FROM doc_packages dp '
            'JOIN construction_stages cs ON dp.stage_id=cs.id '
            'JOIN objects o ON cs.object_id=o.id '
            'LEFT JOIN organizations org ON dp.contractor_id=org.id '
            'ORDER BY dp.created_at DESC')
        for row_i, p in enumerate(packages, 4):
            steps = query_db('SELECT role, status FROM approval_steps WHERE package_id=? ORDER BY step_order', (p['id'],))
            step_map = {s['role']: step_st.get(s['status'],'') for s in steps}
            vals = [p['id'], p['sub_name'], p['obj_name'], p['stage_name'],
                    p['con_name'] or '', pkg_st.get(p['status'],''),
                    step_map.get('inspector',''), step_map.get('foreman',''),
                    step_map.get('pto',''), step_map.get('manager',''),
                    step_map.get('accountant',''),
                    (p['submitted_at'] or '')[:10]]
            for col, val in enumerate(vals, 1):
                ws4.cell(row_i, col, val).border = border4

        # ═══ Лист 5: Заявки на материал ═══
        ws5 = wb.create_sheet('Заявки на материал')
        headers5 = ['#', 'Объект', 'Этап', 'Подрядчик', 'Статус', 'Позиций',
                     'Создана', 'Завершена']
        border5 = _style_sheet(ws5, 'Заявки на давальческий материал', headers5)
        for cw, w in [(1,5),(2,20),(3,18),(4,18),(5,14),(6,8),(7,12),(8,12)]:
            ws5.column_dimensions[chr(64+cw)].width = w

        mr_st = {'submitted':'Отправлена','returned':'Возвращена','approved':'Одобрена',
                 'processing':'В обработке','completed':'Завершена'}
        mrs = query_db(
            'SELECT mr.*, cs.name as stage_name, o.name as obj_name, org.name as con_name, '
            '(SELECT COUNT(*) FROM material_request_items WHERE request_id=mr.id) as items_count '
            'FROM material_requests mr '
            'JOIN construction_stages cs ON mr.stage_id=cs.id '
            'JOIN objects o ON cs.object_id=o.id '
            'LEFT JOIN organizations org ON mr.contractor_id=org.id '
            'ORDER BY mr.created_at DESC')
        for row_i, m in enumerate(mrs, 4):
            vals = [m['id'], m['obj_name'], m['stage_name'], m['con_name'] or '',
                    mr_st.get(m['status'],''), m['items_count'],
                    (m['created_at'] or '')[:10], (m['completed_at'] or '')[:10] if m['completed_at'] else '']
            for col, val in enumerate(vals, 1):
                ws5.cell(row_i, col, val).border = border5

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='ШТАБ_Сводка.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
