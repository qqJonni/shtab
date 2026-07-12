"""График производства работ (ГПР): Гант, план-факт, baseline, вехи."""
import json
from datetime import date, timedelta

from flask import render_template, redirect, url_for, request, flash, abort
from flask_login import login_required, current_user

from db import query_db, execute_db
from helpers import assert_object_access, role_required

RISK_AHEAD_DAYS = 7   # «под риском»: плановый финиш ближе N дней, работа не завершена


def _risk(status, plan_end, today, done_statuses):
    """Цвет полосы: done / suspended / overdue / risk / on_track."""
    if status in done_statuses:
        return 'done'
    if status == 'suspended':
        return 'suspended'
    if plan_end:
        if plan_end < today:
            return 'overdue'
        if plan_end <= (date.fromisoformat(today) + timedelta(days=RISK_AHEAD_DAYS)).isoformat():
            return 'risk'
    return 'on_track'


def _days(a, b):
    """b − a в днях (ISO-строки)."""
    return (date.fromisoformat(b) - date.fromisoformat(a)).days


def _deviation(plan_end, actual_end, finished, today):
    """Отклонение по срокам в днях: >0 опережение, <0 отставание, None — нет данных.

    Завершено: plan_end − actual_end.
    В работе:  если plan_end уже прошёл — накопленное отставание −(today − plan_end),
               иначе 0 (идёт по графику, судить рано)."""
    if not plan_end:
        return None
    if finished and actual_end:
        return _days(actual_end, plan_end)
    if not finished:
        return _days(today, plan_end) if plan_end < today else 0
    return None


def _closed_volumes(object_id):
    """Закрытые объёмы по подэтапам объекта из completed-пакетов.
    qty IS NULL в строке пакета = полное закрытие подэтапа."""
    rows = query_db(
        "SELECT pi.substage_id, "
        "SUM(pi.qty) as qty_sum, COUNT(*) FILTER (WHERE pi.qty IS NULL) as full_cnt "
        "FROM package_items pi "
        "JOIN doc_packages dp ON pi.package_id = dp.id "
        "JOIN construction_stages cs ON dp.stage_id = cs.id "
        "WHERE dp.status = 'completed' AND cs.object_id = ? "
        "GROUP BY pi.substage_id", (object_id,))
    return {r['substage_id']: r for r in rows}


def get_s_curve(object_id, rng_start, rng_end, today):
    """S-кривая освоения: недельные точки, кумулятив в ₽.

    План: стоимость каждого подэтапа равномерно распределяется по его
    плановому периоду (plan_start подэтапа, fallback — plan_start этапа;
    без дат — скачком на плановый финиш).
    Факт: суммы принятых пакетов КС на их completed_at + для завершённых
    подэтапов без пакетов (или закрытых не полностью) — остаток стоимости
    на actual_end_date. Факт обрезается по сегодня."""
    subs = query_db(
        'SELECT ss.*, cs.plan_start_date as st_plan_start '
        'FROM substages ss JOIN construction_stages cs ON ss.stage_id = cs.id '
        'WHERE cs.object_id = ?', (object_id,))
    pkg_events = query_db(
        "SELECT LEFT(dp.completed_at, 10) as d, SUM(pi.amount) as amt, "
        "       pi.substage_id "
        "FROM package_items pi JOIN doc_packages dp ON pi.package_id = dp.id "
        "JOIN construction_stages cs ON dp.stage_id = cs.id "
        "WHERE dp.status = 'completed' AND dp.completed_at IS NOT NULL AND cs.object_id = ? "
        "GROUP BY LEFT(dp.completed_at, 10), pi.substage_id", (object_id,))

    d0 = date.fromisoformat(rng_start)
    d1 = date.fromisoformat(rng_end)
    n_days = (d1 - d0).days + 1
    plan_daily = [0.0] * n_days
    fact_daily = [0.0] * n_days

    def day_idx(iso):
        return min(max((date.fromisoformat(iso) - d0).days, 0), n_days - 1)

    closed_by_sub = {}
    for e in pkg_events:
        closed_by_sub[e['substage_id']] = closed_by_sub.get(e['substage_id'], 0.0) + float(e['amt'] or 0)
        fact_daily[day_idx(e['d'])] += float(e['amt'] or 0)

    for x in subs:
        cost = float(x['total_price'] or 0)
        if cost <= 0:
            continue
        # план
        p_start = x['plan_start_date'] or x['st_plan_start']
        p_end = x['plan_end_date']
        if p_start and p_end and p_start <= p_end:
            i0, i1 = day_idx(p_start), day_idx(p_end)
            per_day = cost / (i1 - i0 + 1)
            for i in range(i0, i1 + 1):
                plan_daily[i] += per_day
        elif p_end:
            plan_daily[day_idx(p_end)] += cost
        # факт: остаток стоимости завершённого подэтапа, не покрытый пакетами
        if x['status'] in ('done', 'closed', 'approved'):
            rest = max(cost - closed_by_sub.get(x['id'], 0.0), 0.0)
            fin = x['actual_end_date'] or (x['completed_at'] or '')[:10] or x['plan_end_date']
            if rest > 0 and fin:
                fact_daily[day_idx(fin)] += rest

    # кумулятив + недельные точки
    labels, plan, fact = [], [], []
    cum_p = cum_f = 0.0
    today_d = date.fromisoformat(today)
    for i in range(n_days):
        cum_p += plan_daily[i]
        cum_f += fact_daily[i]
        d = d0 + timedelta(days=i)
        if d.weekday() == 0 or i == n_days - 1 or i == 0:
            labels.append(d.isoformat())
            plan.append(round(cum_p))
            fact.append(round(cum_f) if d <= today_d else None)

    return {'labels': labels, 'plan': plan, 'fact': fact,
            'total': round(sum(plan_daily))}


def _active_baseline(object_id):
    """Последний утверждённый baseline объекта: (meta, {'stage:ID'|'sub:ID': {...}})."""
    row = query_db(
        'SELECT * FROM schedule_baselines WHERE object_id = ? ORDER BY id DESC LIMIT 1',
        (object_id,), one=True)
    if not row or not row['data_json']:
        return None, {}
    try:
        return dict(row), json.loads(row['data_json'])
    except (ValueError, TypeError):
        return dict(row), {}


def _baseline_dev(item_key, baseline, ref_end):
    """Отклонение от baseline в днях: baseline_end − ref_end.
    ref_end — факт-финиш (для завершённых) или текущий плановый.
    >0 идём раньше утверждённого, <0 — уехали позже."""
    b = baseline.get(item_key)
    if not b or not b.get('plan_end') or not ref_end:
        return None
    return _days(ref_end, b['plan_end'])


def get_schedule_data(object_id):
    """Данные Ганта и план-факта по объекту."""
    today = date.today().isoformat()
    bl_meta, baseline = _active_baseline(object_id)
    stages = query_db(
        'SELECT cs.*, org.name as contractor_name '
        'FROM construction_stages cs '
        'LEFT JOIN organizations org ON cs.contractor_id = org.id '
        'WHERE cs.object_id = ? ORDER BY cs.order_num, cs.id', (object_id,))
    closed = _closed_volumes(object_id)

    out = []
    dates = []
    for st in stages:
        s = dict(st)
        subs = [dict(r) for r in query_db(
            'SELECT * FROM substages WHERE stage_id = ? ORDER BY id', (s['id'],))]

        s['risk'] = _risk(s['status'], s['plan_end_date'], today, ('done',))
        # факт-полоса «в работе» тянется до сегодня
        s['bar_actual_end'] = s['actual_end_date'] or (today if s['actual_start_date'] else None)

        sub_rows = []
        st_plan_cost = 0.0   # плановая стоимость этапа
        st_done_cost = 0.0   # «выполненная» стоимость (для взвешенного %)
        for x in subs:
            fin = x['status'] in ('done', 'closed', 'approved')
            volume = float(x['volume']) if x['volume'] is not None else None
            price = float(x['unit_price']) if x['unit_price'] is not None else 0.0
            cost = float(x['total_price']) if x['total_price'] is not None else (volume or 0) * price

            # закрытый объём: из completed-пакетов; полное закрытие = весь объём
            c = closed.get(x['id'])
            closed_qty = float(c['qty_sum']) if c and c['qty_sum'] is not None else 0.0
            if c and c['full_cnt'] and volume is not None:
                closed_qty = volume
            if volume is not None:
                closed_qty = min(closed_qty, volume)

            # % выполнения: по объёму если есть; иначе по статусу
            if volume:
                progress = round(closed_qty / volume * 100)
                if fin:
                    progress = 100
            else:
                progress = 100 if fin else (50 if x['status'] == 'in_progress' else 0)

            st_plan_cost += cost
            st_done_cost += cost * progress / 100

            dev = _deviation(x['plan_end_date'], x['actual_end_date'], fin, today)
            ref_end = x['actual_end_date'] if fin else x['plan_end_date']
            row = {
                'id': x['id'], 'name': x['name'],
                'plan_start': x['plan_start_date'], 'plan_end': x['plan_end_date'],
                'actual_start': x['actual_start_date'],
                'actual_end': x['actual_end_date'] or (today if x['actual_start_date'] and not fin else x['actual_end_date']),
                'volume': volume, 'unit': x['unit'], 'closed_qty': closed_qty if volume else None,
                'progress': progress, 'status': x['status'],
                'deviation': dev,
                'baseline_dev': _baseline_dev(f'sub:{x["id"]}', baseline, ref_end),
                'risk': _risk(x['status'], x['plan_end_date'], today, ('done', 'closed', 'approved')),
            }
            sub_rows.append(row)
            dates += [x['plan_start_date'], x['plan_end_date'], x['actual_start_date'], x['actual_end_date']]

        # % этапа: взвешенный по стоимости; без цен — по числу завершённых
        if st_plan_cost > 0:
            s['progress'] = round(st_done_cost / st_plan_cost * 100)
        else:
            done_cnt = sum(1 for x in sub_rows if x['progress'] == 100)
            s['progress'] = round(done_cnt / len(sub_rows) * 100) if sub_rows else (100 if s['status'] == 'done' else 0)
        if s['status'] == 'done':
            s['progress'] = 100

        dates += [s['plan_start_date'], s['plan_end_date'], s['actual_start_date'], s['bar_actual_end']]
        out.append({
            'id': s['id'], 'name': s['name'], 'contractor': s['contractor_name'],
            'status': s['status'], 'progress': s['progress'], 'risk': s['risk'],
            'plan_start': s['plan_start_date'], 'plan_end': s['plan_end_date'],
            'actual_start': s['actual_start_date'], 'actual_end': s['bar_actual_end'],
            'deviation': _deviation(s['plan_end_date'], s['actual_end_date'], s['status'] == 'done', today),
            'baseline_dev': _baseline_dev(
                f'stage:{s["id"]}', baseline,
                s['actual_end_date'] if s['status'] == 'done' else s['plan_end_date']),
            'plan_cost': st_plan_cost, 'done_cost': st_done_cost,
            'subs': sub_rows,
        })

    dates = [d for d in dates if d]
    if dates:
        rng_start = min(min(dates), today)
        rng_end = max(max(dates), today)
    else:
        rng_start = rng_end = today
    # паддинг диапазона
    rng_start = (date.fromisoformat(rng_start) - timedelta(days=7)).isoformat()
    rng_end = (date.fromisoformat(rng_end) + timedelta(days=14)).isoformat()

    # Итоги по объекту
    total_cost = sum(s['plan_cost'] for s in out)
    done_cost = sum(s['done_cost'] for s in out)
    all_rows = out + [x for s in out for x in s['subs']]
    overdue_cnt = sum(1 for s in out for x in s['subs'] if x['risk'] == 'overdue')
    # суммарное отставание: по подэтапам с отрицательным отклонением
    lag_total = sum(-x['deviation'] for s in out for x in s['subs']
                    if x['deviation'] is not None and x['deviation'] < 0)
    totals = {
        'progress': round(done_cost / total_cost * 100) if total_cost else (
            round(sum(s['progress'] for s in out) / len(out)) if out else 0),
        'plan_cost': total_cost, 'done_cost': done_cost,
        'overdue': overdue_cnt, 'lag_days': lag_total,
        'stages_total': len(out),
    }

    milestones = [dict(r) for r in query_db(
        'SELECT * FROM schedule_milestones WHERE object_id = ? ORDER BY order_num, plan_date, id',
        (object_id,))]
    for m in milestones:
        m['overdue'] = (m['status'] == 'pending' and m['plan_date'] and m['plan_date'] < today)

    return {'range': {'start': rng_start, 'end': rng_end, 'today': today},
            'stages': out, 'totals': totals,
            'scurve': get_s_curve(object_id, rng_start, rng_end, today),
            'baseline': {'id': bl_meta['id'], 'name': bl_meta['name'],
                         'created_at': bl_meta['created_at']} if bl_meta else None,
            'milestones': milestones}


def register(app):

    @app.route('/objects/<int:obj_id>/schedule')
    @login_required
    def object_schedule(obj_id):
        obj = query_db('SELECT * FROM objects WHERE id = ?', (obj_id,), one=True)
        if not obj:
            abort(404)
        assert_object_access(current_user, obj_id)

        data = get_schedule_data(obj_id)
        return render_template('schedule/view.html', obj=obj, data=data)

    @app.route('/objects/<int:obj_id>/schedule/baseline', methods=['POST'])
    @login_required
    @role_required('manager', 'admin')
    def schedule_baseline_create(obj_id):
        obj = query_db('SELECT * FROM objects WHERE id = ?', (obj_id,), one=True)
        if not obj:
            abort(404)
        assert_object_access(current_user, obj_id)

        snapshot = {}
        for s in query_db('SELECT id, plan_start_date, plan_end_date FROM construction_stages '
                          'WHERE object_id = ?', (obj_id,)):
            snapshot[f'stage:{s["id"]}'] = {'plan_start': s['plan_start_date'], 'plan_end': s['plan_end_date']}
        for x in query_db('SELECT ss.id, ss.plan_start_date, ss.plan_end_date, ss.volume, ss.total_price '
                          'FROM substages ss JOIN construction_stages cs ON ss.stage_id = cs.id '
                          'WHERE cs.object_id = ?', (obj_id,)):
            snapshot[f'sub:{x["id"]}'] = {
                'plan_start': x['plan_start_date'], 'plan_end': x['plan_end_date'],
                'volume': float(x['volume']) if x['volume'] is not None else None,
                'cost': float(x['total_price']) if x['total_price'] is not None else None}

        n = query_db('SELECT COUNT(*) as c FROM schedule_baselines WHERE object_id = ?',
                     (obj_id,), one=True)['c']
        name = request.form.get('name', '').strip() or f'Версия {n + 1} от {date.today().isoformat()}'
        execute_db('INSERT INTO schedule_baselines (object_id, name, created_by, data_json) '
                   'VALUES (?, ?, ?, ?)',
                   (obj_id, name, current_user.id, json.dumps(snapshot, ensure_ascii=False)))
        flash(f'График утверждён: «{name}». Отклонения теперь считаются от этой версии.', 'success')
        return redirect(url_for('object_schedule', obj_id=obj_id))

    # ─── Вехи ────────────────────────────────────────────────────────────────

    @app.route('/objects/<int:obj_id>/milestones/add', methods=['POST'])
    @login_required
    @role_required('manager', 'pto', 'admin')
    def milestone_add(obj_id):
        assert_object_access(current_user, obj_id)
        name = request.form.get('name', '').strip()
        plan_date = request.form.get('plan_date', '').strip() or None
        if not name:
            flash('Введите название вехи.', 'danger')
        else:
            execute_db('INSERT INTO schedule_milestones (object_id, name, plan_date) VALUES (?, ?, ?)',
                       (obj_id, name, plan_date))
            flash('Веха добавлена.', 'success')
        return redirect(url_for('object_schedule', obj_id=obj_id) + '#milestones')

    def _get_milestone_or_403(m_id):
        m = query_db('SELECT * FROM schedule_milestones WHERE id = ?', (m_id,), one=True)
        if not m:
            abort(404)
        assert_object_access(current_user, m['object_id'])
        return m

    @app.route('/milestones/<int:m_id>/done', methods=['POST'])
    @login_required
    @role_required('manager', 'pto', 'admin')
    def milestone_done(m_id):
        m = _get_milestone_or_403(m_id)
        if m['status'] == 'done':
            execute_db("UPDATE schedule_milestones SET status = 'pending', actual_date = NULL WHERE id = ?", (m_id,))
            flash('Веха возвращена в ожидание.', 'info')
        else:
            execute_db("UPDATE schedule_milestones SET status = 'done', actual_date = ? WHERE id = ?",
                       (request.form.get('actual_date', '').strip() or date.today().isoformat(), m_id))
            flash('Веха отмечена достигнутой.', 'success')
        return redirect(url_for('object_schedule', obj_id=m['object_id']) + '#milestones')

    @app.route('/milestones/<int:m_id>/delete', methods=['POST'])
    @login_required
    @role_required('manager', 'pto', 'admin')
    def milestone_delete(m_id):
        m = _get_milestone_or_403(m_id)
        execute_db('DELETE FROM schedule_milestones WHERE id = ?', (m_id,))
        flash('Веха удалена.', 'success')
        return redirect(url_for('object_schedule', obj_id=m['object_id']) + '#milestones')

    # ─── Экспорт ─────────────────────────────────────────────────────────────

    def _fmt_d(s):
        """ISO → ДД.ММ.ГГГГ"""
        if not s:
            return '—'
        return f'{s[8:10]}.{s[5:7]}.{s[0:4]}'

    def _export_data(obj_id):
        obj = query_db('SELECT * FROM objects WHERE id = ?', (obj_id,), one=True)
        if not obj:
            abort(404)
        assert_object_access(current_user, obj_id)
        return obj, get_schedule_data(obj_id)

    @app.route('/objects/<int:obj_id>/schedule/export/xlsx')
    @login_required
    @role_required('manager', 'pto', 'admin')
    def schedule_export_xlsx(obj_id):
        import io
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

        obj, data = _export_data(obj_id)
        has_bl = bool(data['baseline'])

        wb = Workbook()
        ws = wb.active
        ws.title = 'План-факт'
        thin = Border(*[Side(style='thin', color='CBD5E1')] * 4)
        head_fill = PatternFill('solid', fgColor='1E293B')
        stage_fill = PatternFill('solid', fgColor='F1F5F9')
        red = Font(color='DC2626')
        green = Font(color='059669')

        ws['A1'] = f'График производства работ — {obj["name"]}'
        ws['A1'].font = Font(bold=True, size=14)
        t = data['totals']
        ws['A2'] = (f'Выполнение: {t["progress"]}%  ·  Освоено: {t["done_cost"]:,.0f} ₽ из {t["plan_cost"]:,.0f} ₽  ·  '
                    f'Просрочено работ: {t["overdue"]}  ·  Отставание: {t["lag_days"]} дн.').replace(',', ' ')
        if has_bl:
            ws['A3'] = f'Утверждённый график: {data["baseline"]["name"]} от {_fmt_d(data["baseline"]["created_at"][:10])}'

        headers = ['Работа', 'План начало', 'План конец', 'Факт начало', 'Факт конец',
                   'Объём план', 'Ед.', 'Закрыто', '%', 'Откл., дн.'] + (['От БЛ, дн.'] if has_bl else []) + ['Статус']
        r = 5
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=ci, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = head_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        RISK = {'on_track': 'В срок', 'risk': 'Под риском', 'overdue': 'Просрочено',
                'done': 'Завершено', 'suspended': 'Приостановлено'}

        def _dev_cell(row_i, col_i, val):
            cell = ws.cell(row=row_i, column=col_i, value=(f'{val:+d}' if val is not None else '—'))
            if val is not None and val < 0:
                cell.font = red
            elif val is not None and val > 0:
                cell.font = green
            return cell

        r += 1
        for s in data['stages']:
            vals = [s['name'], _fmt_d(s['plan_start']), _fmt_d(s['plan_end']),
                    _fmt_d(s['actual_start']), _fmt_d(s['actual_end']),
                    '—', '—', '—', f'{s["progress"]}%']
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=ci, value=v)
                cell.font = Font(bold=True)
                cell.fill = stage_fill
                cell.border = thin
            ci = 10
            _dev_cell(r, ci, s['deviation']).fill = stage_fill
            if has_bl:
                ci += 1
                _dev_cell(r, ci, s['baseline_dev']).fill = stage_fill
            cell = ws.cell(row=r, column=ci + 1, value=RISK.get(s['risk'], s['risk']))
            cell.fill = stage_fill
            cell.border = thin
            r += 1
            for x in s['subs']:
                vals = ['    ' + x['name'], _fmt_d(x['plan_start']), _fmt_d(x['plan_end']),
                        _fmt_d(x['actual_start']), _fmt_d(x['actual_end']),
                        x['volume'] if x['volume'] is not None else '—', x['unit'] or '—',
                        round(x['closed_qty'], 1) if x['closed_qty'] is not None else '—',
                        f'{x["progress"]}%']
                for ci, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=ci, value=v)
                    cell.border = thin
                ci = 10
                _dev_cell(r, ci, x['deviation'])
                if has_bl:
                    ci += 1
                    _dev_cell(r, ci, x['baseline_dev'])
                ws.cell(row=r, column=ci + 1, value=RISK.get(x['risk'], x['risk'])).border = thin
                r += 1

        # Вехи
        if data['milestones']:
            r += 1
            ws.cell(row=r, column=1, value='Ключевые вехи').font = Font(bold=True, size=12)
            r += 1
            for m in data['milestones']:
                status = 'Достигнута' if m['status'] == 'done' else ('Просрочена' if m['overdue'] else 'Ожидается')
                ws.cell(row=r, column=1, value='◆ ' + m['name']).border = thin
                ws.cell(row=r, column=2, value=_fmt_d(m['plan_date'])).border = thin
                ws.cell(row=r, column=3, value=_fmt_d(m['actual_date'])).border = thin
                cell = ws.cell(row=r, column=4, value=status)
                cell.border = thin
                if m['overdue']:
                    cell.font = red
                r += 1

        widths = [44, 13, 13, 13, 13, 12, 8, 12, 8, 11] + ([11] if has_bl else []) + [15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=5, column=i).column_letter].width = w
        ws.freeze_panes = 'A6'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'ГПР_{obj["name"][:40]}_{date.today().isoformat()}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.route('/objects/<int:obj_id>/schedule/export/pdf')
    @login_required
    @role_required('manager', 'pto', 'admin')
    def schedule_export_pdf(obj_id):
        import io
        from flask import send_file
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.graphics.shapes import Drawing, String
        from reportlab.graphics.charts.lineplots import LinePlot
        from reportlab.graphics.widgets.markers import makeMarker
        from routes.digest import _register_font

        obj, data = _export_data(obj_id)
        has_bl = bool(data['baseline'])
        mf = _register_font()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=12*mm, rightMargin=12*mm,
                                topMargin=12*mm, bottomMargin=12*mm)
        st_title = ParagraphStyle('t', fontName=mf, fontSize=15, leading=19)
        st_sub = ParagraphStyle('s', fontName=mf, fontSize=9.5, leading=13, textColor=colors.HexColor('#64748B'))
        st_h = ParagraphStyle('h', fontName=mf, fontSize=12, leading=15, spaceBefore=8)
        st_cell = ParagraphStyle('c', fontName=mf, fontSize=8, leading=10)

        story = []
        story.append(Paragraph(f'График производства работ — {obj["name"]}', st_title))
        sub = f'Сформирован {_fmt_d(date.today().isoformat())}'
        if has_bl:
            sub += f' · Утверждённый график: {data["baseline"]["name"]} от {_fmt_d(data["baseline"]["created_at"][:10])}'
        story.append(Paragraph(sub, st_sub))
        story.append(Spacer(1, 6*mm))

        # Итоги
        t = data['totals']
        money = lambda v: f'{v:,.0f} руб.'.replace(',', ' ')
        tot = Table([[f'Выполнение: {t["progress"]}%',
                      f'Освоено: {money(t["done_cost"])} из {money(t["plan_cost"])}',
                      f'Просрочено работ: {t["overdue"]}',
                      f'Отставание: {t["lag_days"]} дн.']],
                    colWidths=[65*mm, 95*mm, 55*mm, 55*mm])
        tot.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F1F5F9')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0, 0), (-1, -1), 6), ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(tot)

        # S-кривая
        sc = data['scurve']
        if sc['labels']:
            story.append(Paragraph('Освоение (S-кривая, %)', st_h))
            d0 = date.fromisoformat(sc['labels'][0])
            xs = [(date.fromisoformat(l) - d0).days for l in sc['labels']]
            total = sc['total'] or 1
            plan_pts = [(x, y / total * 100) for x, y in zip(xs, sc['plan'])]
            fact_pts = [(x, y / total * 100) for x, y in zip(xs, sc['fact']) if y is not None]
            dr = Drawing(600, 150)
            lp = LinePlot()
            lp.x, lp.y, lp.width, lp.height = 40, 15, 540, 125
            lp.data = [plan_pts, fact_pts or [(0, 0)]]
            lp.lines[0].strokeColor = colors.HexColor('#94A3B8')
            lp.lines[0].strokeDashArray = [3, 2]
            lp.lines[1].strokeColor = colors.HexColor('#3B82F6')
            lp.lines[1].strokeWidth = 1.6
            lp.yValueAxis.valueMin, lp.yValueAxis.valueMax = 0, 100
            lp.yValueAxis.valueStep = 25
            lp.yValueAxis.labels.fontName = mf
            lp.yValueAxis.labels.fontSize = 7
            lp.xValueAxis.visibleLabels = False
            lp.xValueAxis.visibleTicks = False
            dr.add(lp)
            dr.add(String(40, 3, _fmt_d(sc['labels'][0]), fontName=mf, fontSize=7,
                          fillColor=colors.HexColor('#64748B')))
            dr.add(String(540, 3, _fmt_d(sc['labels'][-1]), fontName=mf, fontSize=7,
                          fillColor=colors.HexColor('#64748B')))
            dr.add(String(430, 138, '— — план      —— факт', fontName=mf, fontSize=8,
                          fillColor=colors.HexColor('#475569')))
            story.append(dr)

        # Вехи
        if data['milestones']:
            story.append(Paragraph('Ключевые вехи', st_h))
            rows = [['Веха', 'План', 'Факт', 'Статус']]
            for m in data['milestones']:
                rows.append([m['name'], _fmt_d(m['plan_date']), _fmt_d(m['actual_date']),
                             'Достигнута' if m['status'] == 'done' else ('Просрочена' if m['overdue'] else 'Ожидается')])
            mt = Table(rows, colWidths=[120*mm, 35*mm, 35*mm, 40*mm])
            mt.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 8.5),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
                ('TOPPADDING', (0, 0), (-1, -1), 3), ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            story.append(mt)

        # План-факт
        story.append(Paragraph('План-факт по работам', st_h))
        head = ['Работа', 'План', 'Факт', 'Объём', 'Закрыто', '%', 'Откл.'] + (['От БЛ'] if has_bl else []) + ['Статус']
        rows = [head]
        styles_extra = []
        RISK = {'on_track': 'В срок', 'risk': 'Риск', 'overdue': 'Просроч.',
                'done': 'Заверш.', 'suspended': 'Приост.'}
        ri = 1
        for s in data['stages']:
            row = [Paragraph(f'<b>{s["name"]}</b>' + (f'<br/><font color="#64748B" size="7">{s["contractor"]}</font>' if s['contractor'] else ''), st_cell),
                   f'{_fmt_d(s["plan_start"])} – {_fmt_d(s["plan_end"])}',
                   f'{_fmt_d(s["actual_start"])} – {_fmt_d(s["actual_end"])}',
                   '—', '—', f'{s["progress"]}%',
                   f'{s["deviation"]:+d}' if s['deviation'] is not None else '—']
            if has_bl:
                row.append(f'{s["baseline_dev"]:+d}' if s['baseline_dev'] is not None else '—')
            row.append(RISK.get(s['risk'], s['risk']))
            rows.append(row)
            styles_extra.append(('BACKGROUND', (0, ri), (-1, ri), colors.HexColor('#F1F5F9')))
            ri += 1
            for x in s['subs']:
                vol = f'{x["volume"]:,.1f} {x["unit"] or ""}'.replace(',', ' ') if x['volume'] is not None else '—'
                clo = f'{x["closed_qty"]:,.1f}'.replace(',', ' ') if x['closed_qty'] is not None else '—'
                row = [Paragraph('&nbsp;&nbsp;&nbsp;' + x['name'], st_cell),
                       f'{_fmt_d(x["plan_start"])} – {_fmt_d(x["plan_end"])}',
                       f'{_fmt_d(x["actual_start"])} – {_fmt_d(x["actual_end"])}',
                       vol, clo, f'{x["progress"]}%',
                       f'{x["deviation"]:+d}' if x['deviation'] is not None else '—']
                if has_bl:
                    row.append(f'{x["baseline_dev"]:+d}' if x['baseline_dev'] is not None else '—')
                row.append(RISK.get(x['risk'], x['risk']))
                rows.append(row)
                if x['risk'] == 'overdue':
                    styles_extra.append(('TEXTCOLOR', (-1, ri), (-1, ri), colors.HexColor('#DC2626')))
                ri += 1

        cw = [72*mm, 42*mm, 42*mm, 22*mm, 18*mm, 12*mm, 14*mm] + ([14*mm] if has_bl else []) + [18*mm]
        pf = Table(rows, colWidths=cw, repeatRows=1)
        pf.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 2.5), ('BOTTOMPADDING', (0, 0), (-1, -1), 2.5),
        ] + styles_extra))
        story.append(pf)

        doc.build(story)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'ГПР_{obj["name"][:40]}_{date.today().isoformat()}.pdf',
                         mimetype='application/pdf')
