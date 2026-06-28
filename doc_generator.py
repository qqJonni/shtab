import os
import json
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config


def generate_document(doc_type, data, package_id, fmt='xlsx'):
    generators = {
        'ks2': _generate_ks2,
        'ks3': _generate_ks3,
        'invoice': _generate_invoice,
        'raw_material_report': _generate_raw_material_report,
    }
    gen = generators.get(doc_type)
    if not gen:
        return None

    if fmt == 'pdf' and doc_type == 'ks2':
        return _generate_ks2_pdf(data, package_id)
    if fmt == 'pdf' and doc_type == 'ks3':
        return _generate_ks3_pdf(data, package_id)
    return gen(data, package_id)


def _ensure_dir(package_id):
    folder = os.path.join(config.PACKAGES_FOLDER, str(package_id))
    os.makedirs(folder, exist_ok=True)
    return folder


def _generate_ks2(data, package_id):
    folder = _ensure_dir(package_id)
    filename = f'ks2_{package_id}.xlsx'
    filepath = os.path.join(folder, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = 'КС-2'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_b = Border(bottom=thin)
    sf = Font(size=8)
    sf9 = Font(size=9)
    bf = Font(bold=True, size=8)
    bf9 = Font(bold=True, size=9)
    bf10 = Font(bold=True, size=10)
    ac = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ar = Alignment(horizontal='right', vertical='center')

    # Ширины колонок: A-H (8 колонок как в образце)
    for c, w in [('A', 7), ('B', 10), ('C', 50), ('D', 16), ('E', 14), ('F', 12), ('G', 14), ('H', 14)]:
        ws.column_dimensions[c].width = w

    # === Правый верхний угол: Унифицированная форма ===
    r = 1
    ws.merge_cells('F1:H1')
    ws['F1'] = 'Унифицированная форма № КС-2'
    ws['F1'].font = Font(size=7)
    ws['F1'].alignment = Alignment(horizontal='right')
    ws.merge_cells('F2:H2')
    ws['F2'] = 'Утверждена постановлением Госкомстата России от 11.11.99 № 100'
    ws['F2'].font = Font(size=6)
    ws['F2'].alignment = Alignment(horizontal='right')

    # Код / Форма по ОКУД / по ОКПО — правая часть
    r = 3
    ws.merge_cells('F3:G3')
    ws['F3'] = 'Форма по ОКУД'
    ws['F3'].font = sf; ws['F3'].alignment = ar
    ws['H3'] = '0322005'
    ws['H3'].font = sf; ws['H3'].border = border; ws['H3'].alignment = ac

    # === Шапка: Инвестор / Заказчик / Подрядчик ===
    r = 4
    ws['A4'] = 'Инвестор'
    ws['A4'].font = bf; ws['A4'].alignment = al

    r = 5
    ws.merge_cells('A5:E5')
    ws['A5'] = data.get('investor_name', '')
    ws['A5'].font = sf9; ws['A5'].alignment = al; ws['A5'].border = border_b
    ws.merge_cells('F5:G5')
    ws['F5'] = 'по ОКПО'
    ws['F5'].font = sf; ws['F5'].alignment = ar
    ws['H5'] = data.get('investor_okpo', '')
    ws['H5'].font = sf; ws['H5'].border = border; ws['H5'].alignment = ac

    r = 6
    ws['A6'] = 'Заказчик (Генподрядчик)'
    ws['A6'].font = bf; ws['A6'].alignment = al

    r = 7
    ws.merge_cells('A7:E7')
    cust_line = f"{data.get('customer_name', '')}, ИНН {data.get('customer_inn', '')}, КПП {data.get('customer_kpp', '')}"
    if data.get('customer_phone'):
        cust_line += f", тел. {data.get('customer_phone', '')}"
    ws['A7'] = cust_line
    ws['A7'].font = sf; ws['A7'].alignment = al; ws['A7'].border = border_b
    ws.merge_cells('F7:G7')
    ws['F7'] = 'по ОКПО'
    ws['F7'].font = sf; ws['F7'].alignment = ar
    ws['H7'] = data.get('customer_okpo', '')
    ws['H7'].font = sf; ws['H7'].border = border; ws['H7'].alignment = ac

    r = 8
    ws['A8'] = 'Подрядчик (Субподрядчик)'
    ws['A8'].font = bf; ws['A8'].alignment = al

    r = 9
    ws.merge_cells('A9:E9')
    con_line = f"{data.get('contractor_name', '')}, ИНН {data.get('contractor_inn', '')}"
    if data.get('contractor_phone'):
        con_line += f", тел. {data.get('contractor_phone', '')}"
    ws['A9'] = con_line
    ws['A9'].font = sf; ws['A9'].alignment = al; ws['A9'].border = border_b
    ws.merge_cells('F9:G9')
    ws['F9'] = 'по ОКПО'
    ws['F9'].font = sf; ws['F9'].alignment = ar
    ws['H9'] = data.get('contractor_okpo', '')
    ws['H9'].font = sf; ws['H9'].border = border; ws['H9'].alignment = ac

    # Стройка
    ws['A10'] = 'Стройка'
    ws['A10'].font = bf
    ws.merge_cells('B10:H10')
    ws['B10'] = data.get('construction_name', '')
    ws['B10'].font = sf; ws['B10'].border = border_b

    # Объект
    ws['A11'] = 'Объект'
    ws['A11'].font = bf
    ws.merge_cells('B11:H11')
    ws['B11'] = data.get('object_name', '')
    ws['B11'].font = sf; ws['B11'].border = border_b

    # Договор подряда (контракт) — правая часть
    ws.merge_cells('E12:F12')
    ws['E12'] = 'Договор подряда (контракт)'
    ws['E12'].font = sf; ws['E12'].alignment = ar
    ws['G12'] = 'номер'
    ws['G12'].font = sf; ws['G12'].alignment = ac
    ws['H12'] = data.get('contract_number', '')
    ws['H12'].font = bf; ws['H12'].border = border; ws['H12'].alignment = ac

    ws['G13'] = 'дата'
    ws['G13'].font = sf; ws['G13'].alignment = ac
    ws['H13'] = data.get('contract_date', '')
    ws['H13'].font = sf; ws['H13'].border = border; ws['H13'].alignment = ac

    # Номер документа / Дата составления / Отчётный период
    ws.merge_cells('D14:E14')
    ws['D14'] = 'Номер документа'
    ws['D14'].font = sf; ws['D14'].alignment = ac; ws['D14'].border = border
    ws['F14'] = 'Дата составления'
    ws['F14'].font = sf; ws['F14'].alignment = ac; ws['F14'].border = border
    ws.merge_cells('G14:H14')
    ws['G14'] = 'Отчётный период'
    ws['G14'].font = sf; ws['G14'].alignment = ac; ws['G14'].border = border

    ws.merge_cells('G15:G15')
    ws['G15'] = 'с'
    ws['G15'].font = sf; ws['G15'].alignment = ac; ws['G15'].border = border
    ws['H15'] = 'по'
    ws['H15'].font = sf; ws['H15'].alignment = ac; ws['H15'].border = border

    # Значения: АКТ номер, дата, период
    ws.merge_cells('A16:C16')
    ws['A16'] = 'АКТ'
    ws['A16'].font = bf10; ws['A16'].alignment = Alignment(horizontal='right', vertical='center')
    ws.merge_cells('D16:E16')
    ws['D16'] = data.get('act_number', '')
    ws['D16'].font = bf9; ws['D16'].alignment = ac; ws['D16'].border = border
    ws['F16'] = data.get('act_date', '')
    ws['F16'].font = sf9; ws['F16'].alignment = ac; ws['F16'].border = border
    ws['G16'] = data.get('period_from', '')
    ws['G16'].font = sf; ws['G16'].alignment = ac; ws['G16'].border = border
    ws['H16'] = data.get('period_to', '')
    ws['H16'].font = sf; ws['H16'].alignment = ac; ws['H16'].border = border

    ws.merge_cells('A17:H17')
    ws['A17'] = 'О ПРИЕМКЕ ВЫПОЛНЕННЫХ РАБОТ'
    ws['A17'].font = bf10; ws['A17'].alignment = Alignment(horizontal='center')

    # Сметная стоимость
    ws.merge_cells('A18:F18')
    ws['A18'] = f"Сметная (договорная) стоимость в соответствии с договором подряда (субподряда)"
    ws['A18'].font = sf
    ws.merge_cells('G18:H18')
    smeta = data.get('smeta_cost', '')
    ws['G18'] = f"{smeta} руб." if smeta else 'руб.'
    ws['G18'].font = sf; ws['G18'].alignment = ar

    # === ЗАГОЛОВКИ ТАБЛИЦЫ ===
    r = 20
    # Строка 20-21: двухуровневый заголовок
    ws.merge_cells(f'A{r}:A{r+1}')
    _c(ws, r, 1, 'Номер\nпо\nпорядку', bf, border, ac)
    ws.merge_cells(f'B{r}:B{r+1}')
    _c(ws, r, 2, 'Номер\nпозиции по\nсмете', bf, border, ac)
    ws.merge_cells(f'C{r}:C{r+1}')
    _c(ws, r, 3, 'Наименование работ', bf, border, ac)
    ws.merge_cells(f'D{r}:D{r+1}')
    _c(ws, r, 4, 'Номер единичной\nрасценки', bf, border, ac)
    ws.merge_cells(f'E{r}:E{r+1}')
    _c(ws, r, 5, 'Единица\nизмерения', bf, border, ac)
    ws.merge_cells(f'F{r}:H{r}')
    _c(ws, r, 6, 'Выполнено работ', bf, border, ac)
    _c(ws, r+1, 6, 'количество', bf, border, ac)
    _c(ws, r+1, 7, 'цена за единицу,\nруб.', bf, border, ac)
    _c(ws, r+1, 8, 'стоимость,\nруб.', bf, border, ac)

    # Номера столбцов
    r = 22
    for i in range(1, 9):
        _c(ws, r, i, str(i), sf, border, ac)

    # === ПОЗИЦИИ ===
    r = 23
    items = data.get('items', [])
    total = 0
    for idx, item in enumerate(items, 1):
        qty = _to_float(item.get('quantity', 0))
        price = _to_float(item.get('price', 0))
        amount = round(qty * price, 2)
        total += amount

        _c(ws, r, 1, idx, sf, border, ac)
        _c(ws, r, 2, item.get('smeta_num', ''), sf, border, ac)
        c = _c(ws, r, 3, item.get('name', ''), sf, border, al)
        _c(ws, r, 4, item.get('rate_num', ''), sf, border, ac)
        _c(ws, r, 5, item.get('unit', ''), sf, border, ac)
        _c(ws, r, 6, qty, sf, border, Alignment(horizontal='right', vertical='center'))
        ws.cell(r, 6).number_format = '#,##0.00'
        _c(ws, r, 7, price, sf, border, ar)
        ws.cell(r, 7).number_format = '#,##0.00'
        _c(ws, r, 8, amount, sf, border, ar)
        ws.cell(r, 8).number_format = '#,##0.00'
        r += 1

    # Итого
    ws.merge_cells(f'A{r}:E{r}')
    _c(ws, r, 1, 'Итого', bf, border, ar)
    for col in range(2, 6):
        ws.cell(r, col).border = border
    _c(ws, r, 6, '', sf, border, ac)
    _c(ws, r, 7, 'X', bf, border, ac)
    _c(ws, r, 8, total, bf, border, ar)
    ws.cell(r, 8).number_format = '#,##0.00'
    r += 1

    # Всего по акту
    ws.merge_cells(f'A{r}:E{r}')
    _c(ws, r, 1, 'Всего по акту', bf, border, ar)
    for col in range(2, 6):
        ws.cell(r, col).border = border
    _c(ws, r, 6, '', sf, border, ac)
    _c(ws, r, 7, 'X', bf, border, ac)
    _c(ws, r, 8, total, bf, border, ar)
    ws.cell(r, 8).number_format = '#,##0.00'
    r += 2

    # === ПОДПИСИ ===
    con_pos = data.get('contractor_rep_position', 'Директор')
    con_name = data.get('contractor_rep_name', '')
    cust_text = data.get('customer_rep_position', '')
    cust_name = data.get('customer_rep_name', '')

    ws['A{}'.format(r)] = 'Сдал'
    ws['A{}'.format(r)].font = bf9
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'] = con_pos
    ws[f'B{r}'].font = sf; ws[f'B{r}'].border = border_b; ws[f'B{r}'].alignment = ac
    ws[f'D{r}'] = '(подпись)'
    ws[f'D{r}'].font = Font(size=7); ws[f'D{r}'].alignment = ac
    ws.merge_cells(f'G{r}:H{r}')
    ws[f'G{r}'] = con_name
    ws[f'G{r}'].font = bf9; ws[f'G{r}'].border = border_b; ws[f'G{r}'].alignment = ac

    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'] = '(должность)'
    ws[f'B{r}'].font = Font(size=7); ws[f'B{r}'].alignment = ac
    ws.merge_cells(f'G{r}:H{r}')
    ws[f'G{r}'] = '(расшифровка подписи)'
    ws[f'G{r}'].font = Font(size=7); ws[f'G{r}'].alignment = ac

    r += 1
    ws[f'B{r}'] = 'М.П.'
    ws[f'B{r}'].font = sf

    r += 2
    ws['A{}'.format(r)] = 'Принял'
    ws['A{}'.format(r)].font = bf9
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'] = cust_text
    ws[f'B{r}'].font = sf; ws[f'B{r}'].border = border_b; ws[f'B{r}'].alignment = al
    ws[f'D{r}'] = '(подпись)'
    ws[f'D{r}'].font = Font(size=7); ws[f'D{r}'].alignment = ac
    ws.merge_cells(f'G{r}:H{r}')
    ws[f'G{r}'] = cust_name
    ws[f'G{r}'].font = bf9; ws[f'G{r}'].border = border_b; ws[f'G{r}'].alignment = ac

    r += 1
    ws.merge_cells(f'B{r}:C{r}')
    ws[f'B{r}'] = '(должность)'
    ws[f'B{r}'].font = Font(size=7); ws[f'B{r}'].alignment = ac
    ws.merge_cells(f'G{r}:H{r}')
    ws[f'G{r}'] = '(расшифровка подписи)'
    ws[f'G{r}'].font = Font(size=7); ws[f'G{r}'].alignment = ac

    r += 1
    ws[f'B{r}'] = 'М.П.'
    ws[f'B{r}'].font = sf

    wb.save(filepath)
    return filename


def _c(ws, row, col, value, font, border, alignment):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.border = border
    cell.alignment = alignment
    return cell


def _generate_ks3(data, package_id):
    folder = _ensure_dir(package_id)
    filename = f'ks3_{package_id}.xlsx'
    filepath = os.path.join(folder, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = 'КС-3'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    lf = Font(bold=True, size=9)
    sf = Font(size=9)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 32
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    ws.merge_cells('A1:F1')
    ws['A1'] = 'СПРАВКА О СТОИМОСТИ ВЫПОЛНЕННЫХ РАБОТ И ЗАТРАТ'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    doc_num = data.get('doc_number', '')
    doc_date = data.get('doc_date', '')
    ws.merge_cells('A2:F2')
    ws['A2'] = f'№ {doc_num} от {doc_date}'
    ws['A2'].font = Font(size=10)
    ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    period = ''
    if data.get('period_from') or data.get('period_to'):
        period = f"с {data.get('period_from', '')} по {data.get('period_to', '')}"

    header_lines = [
        ('Заказчик (Генподрядчик):', f"{data.get('customer_name', '')}  ИНН {data.get('customer_inn', '')}  КПП {data.get('customer_kpp', '')}"),
        ('', f"Адрес: {data.get('customer_address', '')}  ОКПО: {data.get('customer_okpo', '')}"),
        ('Подрядчик (Субподрядчик):', f"{data.get('contractor_name', '')}  ИНН {data.get('contractor_inn', '')}  КПП {data.get('contractor_kpp', '')}"),
        ('', f"Адрес: {data.get('contractor_address', '')}  ОКПО: {data.get('contractor_okpo', '')}"),
        ('Стройка:', data.get('construction_name', '')),
        ('Адрес стройки:', data.get('construction_address', '')),
        ('Договор №:', f"{data.get('contract_number', '')} от {data.get('contract_date', '')}"),
        ('Сметная стоимость:', f"{data.get('smeta_cost', '')} руб."),
        ('Отчётный период:', period),
    ]
    for label, val in header_lines:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = lf
        ws.merge_cells(f'B{row}:F{row}')
        ws[f'B{row}'] = val
        ws[f'B{row}'].font = sf
        row += 1

    row += 1
    for i, h in enumerate(['№ п/п', 'Код', 'Наименование работ и затрат',
                            'Стоимость с начала работ (₽)', 'Стоимость с начала года (₽)',
                            'В т.ч. за отчётный период (₽)']):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = Font(bold=True, size=9)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row += 1
    items = data.get('items', [])
    total_cumulative = 0
    total_year = 0
    total_period = 0
    for idx, item in enumerate(items, 1):
        cum = _to_float(item.get('cumulative', 0))
        year = _to_float(item.get('year_amount', 0))
        period = _to_float(item.get('period_amount', 0))
        total_cumulative += cum
        total_year += year
        total_period += period
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item.get('code', '')).border = border
        ws.cell(row=row, column=3, value=item.get('name', '')).border = border
        c = ws.cell(row=row, column=4, value=cum); c.border = border; c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=5, value=year); c.border = border; c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=6, value=period); c.border = border; c.number_format = '#,##0.00'
        row += 1

    hf = Font(bold=True, size=10)
    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'Итого:'
    ws[f'A{row}'].font = hf; ws[f'A{row}'].alignment = Alignment(horizontal='right')
    for col, val in [(4, total_cumulative), (5, total_year), (6, total_period)]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = hf; c.border = border; c.number_format = '#,##0.00'

    vat_mode = data.get('vat_mode', 'none')
    vat_rate = _to_float(data.get('vat_rate', 20))
    row += 1
    if vat_mode == 'on_top':
        vat_p = total_period * vat_rate / 100
        grand_p = total_period + vat_p
    elif vat_mode == 'included':
        vat_p = total_period * vat_rate / (100 + vat_rate)
        grand_p = total_period
    else:
        vat_p = 0
        grand_p = total_period

    if vat_mode != 'none':
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = f'НДС {int(vat_rate)}%:'
        ws[f'A{row}'].alignment = Alignment(horizontal='right')
        c = ws.cell(row=row, column=6, value=round(vat_p, 2)); c.number_format = '#,##0.00'
        row += 1

    ws.merge_cells(f'A{row}:C{row}')
    ws[f'A{row}'] = 'Всего с НДС:' if vat_mode != 'none' else 'Всего:'
    ws[f'A{row}'].font = Font(bold=True, size=11)
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    c = ws.cell(row=row, column=6, value=round(grand_p, 2))
    c.font = Font(bold=True, size=11); c.number_format = '#,##0.00'

    row += 2
    ws[f'A{row}'] = 'Заказчик (Генподрядчик):'
    ws[f'D{row}'] = 'Подрядчик (Субподрядчик):'

    wb.save(filepath)
    return filename


def _generate_invoice(data, package_id):
    folder = _ensure_dir(package_id)
    filename = f'invoice_{package_id}.xlsx'
    filepath = os.path.join(folder, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Счёт-фактура'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    lf = Font(bold=True, size=9)
    sf = Font(size=9)

    for c, w in [('A', 5), ('B', 30), ('C', 7), ('D', 10), ('E', 12),
                 ('F', 14), ('G', 8), ('H', 12), ('I', 14)]:
        ws.column_dimensions[c].width = w

    doc_num = data.get('doc_number', '')
    doc_date = data.get('doc_date', '')
    ws.merge_cells('A1:I1')
    ws['A1'] = f'СЧЁТ-ФАКТУРА № {doc_num} от {doc_date}'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center')

    corr = data.get('correction_number', '')
    if corr:
        ws.merge_cells('A2:I2')
        ws['A2'] = f'Исправление № {corr} от {data.get("correction_date", "")}'
        ws['A2'].font = sf
        ws['A2'].alignment = Alignment(horizontal='center')

    row = 4
    header_lines = [
        ('Продавец:', f"{data.get('contractor_name', '')}"),
        ('Адрес:', data.get('contractor_address', '')),
        ('ИНН/КПП продавца:', f"{data.get('contractor_inn', '')}/{data.get('contractor_kpp', '')}"),
        ('Грузоотправитель:', data.get('shipper', '') or 'он же'),
        ('Грузополучатель:', data.get('consignee', '') or 'он же'),
        ('К платёжно-расчётному документу:', data.get('payment_doc', '') or '—'),
        ('Покупатель:', data.get('customer_name', '')),
        ('Адрес:', data.get('customer_address', '')),
        ('ИНН/КПП покупателя:', f"{data.get('customer_inn', '')}/{data.get('customer_kpp', '')}"),
        ('Валюта:', data.get('currency', 'Российский рубль, 643')),
    ]
    for label, val in header_lines:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = lf
        ws.merge_cells(f'B{row}:I{row}')
        ws[f'B{row}'] = val
        ws[f'B{row}'].font = sf
        row += 1

    row += 1
    headers = ['№', 'Наименование товара\n(работ, услуг)', 'Ед.', 'Кол-во',
               'Цена (₽)', 'Стоимость\nбез налога', 'Ставка\nНДС',
               'Сумма\nналога', 'Стоимость\nс налогом']
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = Font(bold=True, size=8)
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row += 1
    items = data.get('items', [])
    vat_mode = data.get('vat_mode', 'on_top')
    vat_rate = _to_float(data.get('vat_rate', 20))
    total_no_vat = 0
    total_vat = 0
    total_with_vat = 0

    for idx, item in enumerate(items, 1):
        qty = _to_float(item.get('quantity', 0))
        price = _to_float(item.get('price', 0))
        base = qty * price

        if vat_mode == 'on_top':
            line_vat = base * vat_rate / 100
            amount_no_vat = base
            amount_with_vat = base + line_vat
        elif vat_mode == 'included':
            line_vat = base * vat_rate / (100 + vat_rate)
            amount_no_vat = base - line_vat
            amount_with_vat = base
        else:
            line_vat = 0
            amount_no_vat = base
            amount_with_vat = base

        total_no_vat += amount_no_vat
        total_vat += line_vat
        total_with_vat += amount_with_vat

        vat_label = 'Без НДС' if vat_mode == 'none' else f'{int(vat_rate)}%'

        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=item.get('name', '')).border = border
        ws.cell(row=row, column=3, value=item.get('unit', '')).border = border
        c = ws.cell(row=row, column=4, value=qty); c.border = border; c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=5, value=price); c.border = border; c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=6, value=round(amount_no_vat, 2)); c.border = border; c.number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=vat_label).border = border
        c = ws.cell(row=row, column=8, value=round(line_vat, 2)); c.border = border; c.number_format = '#,##0.00'
        c = ws.cell(row=row, column=9, value=round(amount_with_vat, 2)); c.border = border; c.number_format = '#,##0.00'
        row += 1

    hf = Font(bold=True, size=10)
    ws.merge_cells(f'A{row}:E{row}')
    ws[f'A{row}'] = 'Всего к оплате:'
    ws[f'A{row}'].font = hf
    ws[f'A{row}'].alignment = Alignment(horizontal='right')
    for col, val in [(6, round(total_no_vat, 2)), (8, round(total_vat, 2)), (9, round(total_with_vat, 2))]:
        c = ws.cell(row=row, column=col, value=val)
        c.font = hf; c.border = border; c.number_format = '#,##0.00'

    row += 2
    head = data.get('contractor_head', '')
    accountant = data.get('contractor_accountant', '')
    ws[f'A{row}'] = f'Руководитель организации: __________ {head}'
    ws[f'A{row}'].font = sf
    row += 1
    ws[f'A{row}'] = f'Главный бухгалтер: __________ {accountant}'
    ws[f'A{row}'].font = sf

    wb.save(filepath)
    return filename


def _generate_raw_material_report(data, package_id):
    folder = _ensure_dir(package_id)
    filename = f'raw_material_{package_id}.xlsx'
    filepath = os.path.join(folder, filename)

    BLANK = os.path.join(config.BASE_DIR, 'templates', 'forms', 'otchet_davalchesky_blank.xlsx')

    if os.path.exists(BLANK):
        wb = load_workbook(BLANK)
        ws = wb.active
        return _fill_blank_template(ws, wb, data, filepath, filename)

    # Fallback if blank not found
    wb = Workbook()
    ws = wb.active
    ws.title = 'Давальческий'
    ws['A1'] = 'ОТЧЁТ (бланк не найден — упрощённый формат)'
    items = data.get('items', [])
    row = 3
    for idx, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item.get('name', ''))
        row += 1
    wb.save(filepath)
    return filename


def _fill_blank_template(ws, wb, data, filepath, filename):
    # Blank structure (verified):
    # C2: title (keep)
    # A4:I4 merged — объект (строка 1)
    # A5:I5 merged — объект (строка 2, кадастр)
    # Row 6-13: свободные строки для шапки (не объединены)
    #   6: дата, 8: генподрядчик, 10: подрядчик, 12: договор
    # Row 14: заголовки таблицы
    # Rows 15-35: данные (A:D merged)
    # Row 39: C=Генподрядчик, G=Подрядчик (подписи)
    # Row 51: C, G = __________ /Ф.И.О./

    sf = Font(size=8)

    # Row 4 (merged A4:I4): наименование объекта
    ws['A4'] = f"Объект: {data.get('object_name', '')}"
    ws['A4'].font = sf

    # Row 5 (merged A5:I5): кадастровый номер
    ws['A5'] = f"Кадастровый номер: {data.get('cadastral_number', '')}"
    ws['A5'].font = sf

    # Row 6: дата
    ws['I6'] = data.get('doc_date', '')
    ws['I6'].font = sf

    # Row 8: генподрядчик
    ws['A8'] = f"Генподрядчик: {data.get('customer_name', '')}  ИНН {data.get('customer_inn', '')} КПП {data.get('customer_kpp', '')}"
    ws['A8'].font = sf

    # Row 10: подрядчик
    ws['A10'] = f"Подрядчик: {data.get('contractor_name', '')}  ИНН {data.get('contractor_inn', '')} КПП {data.get('contractor_kpp', '')}"
    ws['A10'].font = sf

    # Row 12: договор
    ws['A12'] = f"Договор: № {data.get('contract_number', '')} от {data.get('contract_date', '')}"
    ws['A12'].font = sf

    # Fill data rows starting at row 15
    items = data.get('items', [])
    DATA_START = 15
    SIGN_ROW_OFFSET = 39 - 35  # signatures are at row 39, data ends at 35 in blank

    # If more items than available rows (15-35 = 21 rows), insert rows
    available = 35 - DATA_START + 1  # 21 rows
    needed = len(items)

    if needed > available:
        extra = needed - available
        ws.insert_rows(DATA_START + available, extra)

    for idx, item in enumerate(items):
        r = DATA_START + idx
        beg = _to_float(item.get('beginning', 0))
        issued = _to_float(item.get('issued', 0))
        certified = _to_float(item.get('certified', 0))
        remainder = beg + issued - certified

        # Unmerge A:D for this row if merged, then re-merge
        for mg in list(ws.merged_cells.ranges):
            if mg.min_row == r and mg.min_col == 1 and mg.max_col == 4:
                ws.unmerge_cells(str(mg))
                break

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=item.get('name', ''))
        c.font = sf

        ws.cell(row=r, column=5, value=item.get('unit', '')).font = sf
        c = ws.cell(row=r, column=6, value=beg); c.font = sf; c.number_format = '#,##0.00'
        c = ws.cell(row=r, column=7, value=issued); c.font = sf; c.number_format = '#,##0.00'
        c = ws.cell(row=r, column=8, value=certified); c.font = sf; c.number_format = '#,##0.00'
        c = ws.cell(row=r, column=9, value=remainder); c.font = sf; c.number_format = '#,##0.00'

    # Fill signatures
    sign_base = DATA_START + max(needed, available) + SIGN_ROW_OFFSET
    # Find the actual signature rows (look for 'Генподрядчик' text)
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=3).value
        if v and 'Генподрядчик' in str(v):
            break
    else:
        r = sign_base

    # Write rep names near signature lines
    cust_rep = data.get('customer_rep', '')
    con_rep = data.get('contractor_rep', '')
    # Find the /Ф.И.О./ row
    for sr in range(r, ws.max_row + 1):
        v = ws.cell(row=sr, column=3).value
        if v and 'Ф.И.О.' in str(v):
            ws.cell(row=sr, column=3, value=f'__________ /{cust_rep}/').font = sf
            break
    for sr in range(r, ws.max_row + 1):
        v = ws.cell(row=sr, column=7).value
        if v and 'Ф.И.О.' in str(v):
            ws.cell(row=sr, column=7, value=f'__________ /{con_rep}/').font = sf
            break

    wb.save(filepath)
    return filename


def _fmt_date(val):
    if not val:
        return ''
    val = str(val).strip()
    if len(val) == 10 and val[4] == '-':
        parts = val.split('-')
        return f'{parts[2]}.{parts[1]}.{parts[0]}'
    return val


def _generate_ks2_pdf(data, package_id):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import reportlab.lib.enums as enums

    folder = _ensure_dir(package_id)
    filename = f'ks2_{package_id}.pdf'
    filepath = os.path.join(folder, filename)

    # Register font - try system fonts
    font_registered = False
    for font_path in [
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/System/Library/Fonts/Helvetica.ttc',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    ]:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('MainFont', font_path))
                font_registered = True
                break
            except Exception:
                continue
    main_font = 'MainFont' if font_registered else 'Helvetica'

    doc = SimpleDocTemplate(filepath, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)

    styles = getSampleStyleSheet()
    s_normal = ParagraphStyle('n', fontName=main_font, fontSize=7, leading=9)
    s_bold = ParagraphStyle('b', fontName=main_font, fontSize=7, leading=9, fontWeight='bold')
    s_center = ParagraphStyle('c', fontName=main_font, fontSize=7, leading=9, alignment=enums.TA_CENTER)
    s_right = ParagraphStyle('r', fontName=main_font, fontSize=7, leading=9, alignment=enums.TA_RIGHT)
    s_title = ParagraphStyle('t', fontName=main_font, fontSize=9, leading=11, alignment=enums.TA_CENTER)
    s_small = ParagraphStyle('sm', fontName=main_font, fontSize=6, leading=7)

    elements = []

    # Header line
    elements.append(Paragraph('Унифицированная форма № КС-2', s_right))
    elements.append(Paragraph('Утверждена постановлением Госкомстата России от 11.11.99 № 100', s_right))
    elements.append(Spacer(1, 3*mm))

    # Parties
    inv = data.get('investor_name', '')
    cust = f"{data.get('customer_name', '')}, ИНН {data.get('customer_inn', '')}, КПП {data.get('customer_kpp', '')}"
    con = f"{data.get('contractor_name', '')}, ИНН {data.get('contractor_inn', '')}"
    if data.get('contractor_phone'):
        con += f", тел. {data.get('contractor_phone')}"

    header_data = [
        ['Инвестор', inv, '', '', 'по ОКПО', data.get('investor_okpo', '')],
        ['Заказчик (Генподрядчик)', cust, '', '', 'по ОКПО', data.get('customer_okpo', '')],
        ['Подрядчик (Субподрядчик)', con, '', '', 'по ОКПО', data.get('contractor_okpo', '')],
        ['Стройка', data.get('construction_name', ''), '', '', '', ''],
        ['Объект', data.get('object_name', ''), '', '', '', ''],
    ]

    ht = Table(header_data, colWidths=[45*mm, 80*mm, 30*mm, 30*mm, 25*mm, 30*mm])
    ht.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), main_font),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('SPAN', (1, 0), (3, 0)), ('SPAN', (1, 1), (3, 1)),
        ('SPAN', (1, 2), (3, 2)), ('SPAN', (1, 3), (3, 3)), ('SPAN', (1, 4), (3, 4)),
        ('LINEBELOW', (1, 0), (3, 0), 0.5, colors.black),
        ('LINEBELOW', (1, 1), (3, 1), 0.5, colors.black),
        ('LINEBELOW', (1, 2), (3, 2), 0.5, colors.black),
        ('LINEBELOW', (1, 3), (3, 3), 0.5, colors.black),
        ('LINEBELOW', (1, 4), (3, 4), 0.5, colors.black),
        ('BOX', (5, 0), (5, 2), 0.5, colors.black),
    ]))
    elements.append(ht)
    elements.append(Spacer(1, 2*mm))

    # Contract + doc info
    contract_num = data.get('contract_number', '')
    contract_date = _fmt_date(data.get('contract_date', ''))
    act_num = data.get('act_number', '')
    act_date = _fmt_date(data.get('act_date', ''))
    period_from = _fmt_date(data.get('period_from', ''))
    period_to = _fmt_date(data.get('period_to', ''))

    info_data = [
        ['', '', '', 'Договор подряда (контракт)', 'номер', contract_num],
        ['', '', '', '', 'дата', contract_date],
        ['', 'Номер документа', 'Дата составления', '', 'Отчётный период', ''],
        ['', '', '', '', 'с', 'по'],
        ['АКТ', act_num, act_date, '', period_from, period_to],
    ]
    it = Table(info_data, colWidths=[40*mm, 40*mm, 40*mm, 40*mm, 40*mm, 40*mm])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), main_font),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 4), (0, 4), 9),
        ('BOX', (1, 4), (1, 4), 0.5, colors.black),
        ('BOX', (2, 4), (2, 4), 0.5, colors.black),
        ('BOX', (4, 4), (4, 4), 0.5, colors.black),
        ('BOX', (5, 4), (5, 4), 0.5, colors.black),
        ('BOX', (5, 0), (5, 1), 0.5, colors.black),
        ('SPAN', (3, 0), (3, 1)),
    ]))
    elements.append(it)
    elements.append(Paragraph('<b>О ПРИЕМКЕ ВЫПОЛНЕННЫХ РАБОТ</b>', s_title))
    elements.append(Spacer(1, 2*mm))

    smeta = data.get('smeta_cost', '')
    elements.append(Paragraph(
        f'Сметная (договорная) стоимость в соответствии с договором подряда (субподряда) &nbsp;&nbsp;&nbsp;&nbsp; {smeta} руб.', s_normal))
    elements.append(Spacer(1, 3*mm))

    # Items table
    items = data.get('items', [])
    total = 0

    tbl_data = [
        ['Номер\nпо\nпорядку', 'Номер\nпозиции\nпо смете', 'Наименование работ', 'Номер единичной\nрасценки',
         'Единица\nизмерения', 'количество', 'цена за единицу,\nруб.', 'стоимость,\nруб.'],
        ['1', '2', '3', '4', '5', '6', '7', '8'],
    ]

    for idx, item in enumerate(items, 1):
        qty = _to_float(item.get('quantity', 0))
        price = _to_float(item.get('price', 0))
        amount = round(qty * price, 2)
        total += amount
        tbl_data.append([
            str(idx), item.get('smeta_num', ''), item.get('name', ''),
            item.get('rate_num', ''), item.get('unit', ''),
            f'{qty:,.2f}'.replace(',', ' '), f'{price:,.2f}'.replace(',', ' '),
            f'{amount:,.2f}'.replace(',', ' '),
        ])

    tbl_data.append(['', '', '', '', 'Итого', '', 'X', f'{total:,.2f}'.replace(',', ' ')])
    tbl_data.append(['', '', '', '', 'Всего по акту', '', 'X', f'{total:,.2f}'.replace(',', ' ')])

    col_w = [12*mm, 14*mm, 75*mm, 22*mm, 18*mm, 22*mm, 22*mm, 25*mm]
    t = Table(tbl_data, colWidths=col_w, repeatRows=2)
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), main_font),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('ALIGN', (5, 2), (7, -1), 'RIGHT'),
        ('ALIGN', (0, 2), (0, -1), 'CENTER'),
        ('ALIGN', (1, 2), (1, -1), 'CENTER'),
        ('ALIGN', (4, 2), (4, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, 1), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 5*mm))

    # Signatures
    con_pos = data.get('contractor_rep_position', '')
    con_name = data.get('contractor_rep_name', '')
    cust_pos = data.get('customer_rep_position', '')
    cust_name = data.get('customer_rep_name', '')

    s_sig = ParagraphStyle('sig', fontName=main_font, fontSize=7, leading=9)

    sig_data = [
        ['Сдал', Paragraph(con_pos, s_sig), '(подпись)', '', con_name],
        ['', '(должность)', '', '', '(расшифровка подписи)'],
        ['', 'М.П.', '', '', ''],
        ['', '', '', '', ''],
        ['Принял', Paragraph(cust_pos, s_sig), '(подпись)', '', cust_name],
        ['', '(должность)', '', '', '(расшифровка подписи)'],
        ['', 'М.П.', '', '', ''],
    ]
    st = Table(sig_data, colWidths=[20*mm, 80*mm, 25*mm, 15*mm, 60*mm])
    st.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), main_font),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEBELOW', (1, 0), (1, 0), 0.5, colors.black),
        ('LINEBELOW', (4, 0), (4, 0), 0.5, colors.black),
        ('LINEBELOW', (1, 4), (1, 4), 0.5, colors.black),
        ('LINEBELOW', (4, 4), (4, 4), 0.5, colors.black),
        ('FONTSIZE', (1, 1), (1, 1), 6),
        ('FONTSIZE', (4, 1), (4, 1), 6),
        ('FONTSIZE', (2, 0), (2, 0), 6),
        ('FONTSIZE', (1, 5), (1, 5), 6),
        ('FONTSIZE', (4, 5), (4, 5), 6),
        ('FONTSIZE', (2, 4), (2, 4), 6),
        ('ALIGN', (1, 1), (1, 1), 'CENTER'),
        ('ALIGN', (4, 1), (4, 1), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'CENTER'),
        ('ALIGN', (1, 5), (1, 5), 'CENTER'),
        ('ALIGN', (4, 5), (4, 5), 'CENTER'),
        ('ALIGN', (2, 4), (2, 4), 'CENTER'),
    ]))
    elements.append(st)

    doc.build(elements)
    return filename


def _generate_ks3_pdf(data, package_id):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import reportlab.lib.enums as enums

    folder = _ensure_dir(package_id)
    filename = f'ks3_{package_id}.pdf'
    filepath = os.path.join(folder, filename)

    font_registered = False
    for font_path in [
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        '/System/Library/Fonts/Helvetica.ttc',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    ]:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont('MainFont', font_path))
                font_registered = True
                break
            except Exception:
                continue
    mf = 'MainFont' if font_registered else 'Helvetica'

    doc = SimpleDocTemplate(filepath, pagesize=A4,
                            leftMargin=12*mm, rightMargin=10*mm,
                            topMargin=8*mm, bottomMargin=8*mm)

    sn = ParagraphStyle('n', fontName=mf, fontSize=7, leading=9)
    sb = ParagraphStyle('b', fontName=mf, fontSize=7, leading=9, spaceAfter=0)
    sc = ParagraphStyle('c', fontName=mf, fontSize=7, leading=9, alignment=enums.TA_CENTER)
    sr = ParagraphStyle('r', fontName=mf, fontSize=7, leading=9, alignment=enums.TA_RIGHT)
    st = ParagraphStyle('t', fontName=mf, fontSize=9, leading=11, alignment=enums.TA_CENTER)
    ss = ParagraphStyle('sig', fontName=mf, fontSize=7, leading=9)

    elements = []

    # Header
    elements.append(Paragraph('Унифицированная форма № КС-3', sr))
    elements.append(Paragraph('Утверждена постановлением Госкомстата России от 11.11.99 № 100', sr))
    elements.append(Spacer(1, 2*mm))

    # Parties
    cust = data.get('customer_name', '')
    if data.get('customer_inn'):
        cust += f", ИНН {data.get('customer_inn', '')}, КПП {data.get('customer_kpp', '')}"
    if data.get('customer_phone'):
        cust += f", тел. {data.get('customer_phone', '')}"

    con = data.get('contractor_name', '')
    if data.get('contractor_inn'):
        con += f", ИНН {data.get('contractor_inn', '')}"
    if data.get('contractor_phone'):
        con += f", тел. {data.get('contractor_phone', '')}"

    hdr = [
        ['Инвестор', '', '', 'по ОКПО', data.get('investor_okpo', '')],
        ['Заказчик (Генподрядчик)', Paragraph(cust, sn), '', 'по ОКПО', data.get('customer_okpo', '')],
        ['Подрядчик (Субподрядчик)', Paragraph(con, sn), '', 'по ОКПО', data.get('contractor_okpo', '')],
        ['Стройка', Paragraph(data.get('construction_name', ''), sn), '', '', ''],
    ]
    ht = Table(hdr, colWidths=[38*mm, 75*mm, 20*mm, 18*mm, 25*mm])
    ht.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('SPAN', (1, 0), (2, 0)), ('SPAN', (1, 1), (2, 1)),
        ('SPAN', (1, 2), (2, 2)), ('SPAN', (1, 3), (2, 3)),
        ('LINEBELOW', (1, 0), (2, 0), 0.5, colors.black),
        ('LINEBELOW', (1, 1), (2, 1), 0.5, colors.black),
        ('LINEBELOW', (1, 2), (2, 2), 0.5, colors.black),
        ('LINEBELOW', (1, 3), (2, 3), 0.5, colors.black),
        ('BOX', (4, 0), (4, 2), 0.5, colors.black),
    ]))
    elements.append(ht)
    elements.append(Spacer(1, 1*mm))

    # Contract + doc info
    contract_num = data.get('contract_number', '')
    contract_date = _fmt_date(data.get('contract_date', ''))
    doc_num = data.get('doc_number', '')
    doc_date = _fmt_date(data.get('doc_date', ''))
    period_from = _fmt_date(data.get('period_from', ''))
    period_to = _fmt_date(data.get('period_to', ''))

    info = [
        ['', '', 'Договор подряда (контракт)', 'номер', contract_num],
        ['', '', '', 'дата', contract_date],
        ['', '', '', 'Вид операции', ''],
        ['', 'Номер документа', 'Дата составления', 'Отчётный период', ''],
        ['', '', '', 'с', 'по'],
        ['СПРАВКА', doc_num, doc_date, period_from, period_to],
    ]
    it = Table(info, colWidths=[30*mm, 35*mm, 40*mm, 35*mm, 35*mm])
    it.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 5), (0, 5), 9),
        ('BOX', (1, 5), (1, 5), 0.5, colors.black),
        ('BOX', (2, 5), (2, 5), 0.5, colors.black),
        ('BOX', (3, 5), (3, 5), 0.5, colors.black),
        ('BOX', (4, 5), (4, 5), 0.5, colors.black),
        ('BOX', (4, 0), (4, 1), 0.5, colors.black),
    ]))
    elements.append(it)
    elements.append(Paragraph('<b>О СТОИМОСТИ ВЫПОЛНЕННЫХ РАБОТ И ЗАТРАТ</b>', st))
    elements.append(Spacer(1, 3*mm))

    # Items table
    items = data.get('items', [])
    total_cum = 0
    total_year = 0
    total_period = 0

    s_wrap = ParagraphStyle('wrap', fontName=mf, fontSize=7, leading=9)

    tbl = [
        ['Номер\nпо\nпорядку', 'Наименование пусковых комплексов, этапов, объектов,\nвидов выполненных работ, оборудования, затрат', 'Код',
         'Стоимость выполненных работ и затрат, руб.', '', ''],
        ['', '', '', 'с начала\nпроведения работ', 'с начала года', 'в том числе за\nотчётный период'],
        ['1', '2', '3', '4', '5', '6'],
    ]

    for idx, item in enumerate(items, 1):
        cum = _to_float(item.get('cumulative', 0))
        year = _to_float(item.get('year_amount', 0))
        period = _to_float(item.get('period_amount', 0))
        total_cum += cum
        total_year += year
        total_period += period
        tbl.append([
            str(idx),
            Paragraph(item.get('name', ''), s_wrap),
            item.get('code', ''),
            f'{cum:,.2f}'.replace(',', ' '),
            f'{year:,.2f}'.replace(',', ' '),
            f'{period:,.2f}'.replace(',', ' '),
        ])

    # Итого / НДС / Всего
    tbl.append(['', '', '', '', 'Итого', f'{total_period:,.2f}'.replace(',', ' ')])

    vat_mode = data.get('vat_mode', 'none')
    vat_rate = _to_float(data.get('vat_rate', 20))
    if vat_mode == 'on_top':
        vat_p = total_period * vat_rate / 100
        grand = total_period + vat_p
    elif vat_mode == 'included':
        vat_p = total_period * vat_rate / (100 + vat_rate)
        grand = total_period
    else:
        vat_p = 0
        grand = total_period

    tbl.append(['', '', '', '', 'Сумма НДС', f'{vat_p:,.2f}'.replace(',', ' ')])
    tbl.append(['', '', '', '', 'Всего с учетом НДС', f'{grand:,.2f}'.replace(',', ' ')])

    cw = [14*mm, 70*mm, 14*mm, 30*mm, 28*mm, 28*mm]
    t = Table(tbl, colWidths=cw, repeatRows=3)
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 2), 'CENTER'), ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('SPAN', (3, 0), (5, 0)),
        ('SPAN', (0, 0), (0, 1)), ('SPAN', (1, 0), (1, 1)), ('SPAN', (2, 0), (2, 1)),
        ('ALIGN', (3, 3), (5, -1), 'RIGHT'),
        ('ALIGN', (0, 3), (0, -1), 'CENTER'),
        ('FONTSIZE', (0, 2), (-1, 2), 6),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 5*mm))

    # Signatures
    cust_pos = data.get('customer_rep_position', '')
    cust_name = data.get('customer_rep_name', '')
    con_pos = data.get('contractor_rep_position', 'Директор')
    con_name = data.get('contractor_rep_name', '')

    sig = [
        ['Заказчик (Генподрядчик)', Paragraph(cust_pos, ss), '(подпись)', '', cust_name],
        ['', '(должность)', '', '', '(расшифровка подписи)'],
        ['', 'М.П.', '', '', ''],
        ['', '', '', '', ''],
        ['Подрядчик (Субподрядчик)', Paragraph(con_pos, ss), '(подпись)', '', con_name],
        ['', '(должность)', '', '', '(расшифровка подписи)'],
        ['', 'М.П.', '', '', ''],
    ]
    st2 = Table(sig, colWidths=[35*mm, 50*mm, 25*mm, 15*mm, 50*mm])
    st2.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), mf), ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LINEBELOW', (1, 0), (1, 0), 0.5, colors.black),
        ('LINEBELOW', (4, 0), (4, 0), 0.5, colors.black),
        ('LINEBELOW', (1, 4), (1, 4), 0.5, colors.black),
        ('LINEBELOW', (4, 4), (4, 4), 0.5, colors.black),
        ('FONTSIZE', (1, 1), (1, 1), 6), ('FONTSIZE', (4, 1), (4, 1), 6),
        ('FONTSIZE', (2, 0), (2, 0), 6), ('FONTSIZE', (2, 4), (2, 4), 6),
        ('FONTSIZE', (1, 5), (1, 5), 6), ('FONTSIZE', (4, 5), (4, 5), 6),
        ('ALIGN', (1, 1), (1, 1), 'CENTER'), ('ALIGN', (4, 1), (4, 1), 'CENTER'),
        ('ALIGN', (2, 0), (2, 0), 'CENTER'), ('ALIGN', (2, 4), (2, 4), 'CENTER'),
        ('ALIGN', (1, 5), (1, 5), 'CENTER'), ('ALIGN', (4, 5), (4, 5), 'CENTER'),
    ]))
    elements.append(st2)

    doc.build(elements)
    return filename


def _to_float(val):
    try:
        return float(val) if val else 0.0
    except (ValueError, TypeError):
        return 0.0
